import io
import os
import traceback
import pandas as pd
from io import StringIO
from datetime import datetime
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VERSAO         = "V3.3"
NOME_CFOP_XLSX = "160314_Tabela_CFOP.xlsx"
NOME_IMP_XLSX  = "Impostos.xlsx"

IMPOSTOS_FALLBACK = {
    'ICMS': 1, 'IPI': 2, 'ISS': 3, 'PIS': 4, 'COFINS': 5,
    'CONTRIBUICAO SOCIAL': 6, 'DIFAL': 8, 'SUBST. TRIBUTARIA': 9,
    'SIMPLES': 10, 'ICMS RETIDO': 11, 'ICMS SUBSTITUTO': 12,
    'PIS NÃO CUMULATIVO': 17, 'COFINS NÃO CUMULATIVO': 19,
    'INSS RETIDO': 26, 'ICMS ANTECIPADO': 27, 'SIMPLES NACIONAL': 44,
    'ICMS IMPORTACAO': 45, 'PIS IMPORTACAO': 133, 'COFINS IMPORTACAO': 134,
    'ICMS DIFERIDO': 116, 'ICMS COMPLEMENTAR': 125,
    'DIFAL NÃO CONTRIBUINTE': 145, 'DIFAL FCP': 146, 'IBS': 183, 'CBS': 184,
}

_ALIQ_TOL = 0.001
ALIQ_PIS_COFINS = {
    'PIS':    [(0.65, _ALIQ_TOL, 4), (1.65, _ALIQ_TOL, 17)],
    'COFINS': [(3.00, _ALIQ_TOL, 5), (7.60, _ALIQ_TOL, 19)],
}

# ==============================
# DE-PARA: CST ENTRADA → CST SAÍDA
# Conforme leiaute Domínio e documentos CST fornecidos.
# CSTs de entrada (50-99) são convertidos para o CST de saída equivalente.
# ==============================
CST_ENTRADA_PARA_SAIDA = {
    '50': '01',  # Aquisição com crédito vinc. a receita tributada → Tributável alíq. básica
    '51': '01',  # Aquisição com crédito vinc. a receita não tributada → Tributável alíq. básica
    '52': '01',  # Aquisição com crédito vinc. a receita de exportação → Tributável alíq. básica
    '53': '01',  # Aquisição com crédito vinc. a receitas tributadas e não tributadas
    '54': '01',
    '55': '01',
    '56': '01',
    '60': '01',  # Crédito presumido → Tributável alíq. básica
    '61': '01',
    '62': '01',
    '63': '01',
    '64': '01',
    '65': '01',
    '66': '01',
    '67': '01',
    '70': '07',  # Aquisição sem direito a crédito → Isenta
    '71': '07',  # Aquisição com isenção → Isenta
    '72': '09',  # Aquisição com suspensão → Suspensão
    '73': '06',  # Aquisição a alíquota zero → Alíquota zero
    '74': '08',  # Aquisição sem incidência → Sem incidência
    '75': '05',  # Aquisição por substituição tributária → Substituição tributária
    '98': '49',  # Outras operações de entrada → Outras operações de saída
    '99': '99',  # Outras operações
}

# ==============================
# CSTs QUE EXIGEM NATUREZA DE RECEITA (obrigatório)
# Fonte: Tabelas 4.3.10, 4.3.13, 4.3.14, 4.3.15, 4.3.16
# ==============================
CSTS_NATUREZA_OBRIGATORIA = {'02', '04', '06', '07', '08', '09'}

# ==============================
# TABELA NATUREZA DE RECEITA — CSTs obrigatórios
# Fonte oficial: CST PIS e COFINS.xlsx (Receita Federal)
# ==============================
NATUREZA_POR_CST = {

    # ── CST 02 — Tabela 4.3.10 (v1.25, 30/03/2026) ──────────────────────
    '02': {
        '001':   'Revenda de combustíveis – Alíquota zero',
        '002':   'Revenda de fármacos e perfumarias – Alíquota zero',
        '003':   'Revenda de veículos, máquinas e autopeças – Alíquota zero',
        '004':   'Revenda de bebidas frias – Alíquota zero (até 30/04/2015)',
        '101': 'Gasolinas, Exceto Gasolina de Aviação',
        '102': 'Óleo Diesel',
        '103': 'Gás Liquefeito de Petróleo – GLP',
        '104': 'Querosene de Aviação',
        '105': 'Correntes Destinadas à Formulação de Gasolinas',
        '106': 'Correntes Destinadas à Formulação de Óleo Diesel',
        '107': 'Nafta Petroquímica Destinada à Formulação de Gasolina ou de Óleo Diesel',
        '108': 'Nafta Petroquímica Destinada à Formulação Exclusivamente de Óleo Diesel',
        '109': 'Biodiesel',
        '112': 'Álcool, Inclusive para Fins Carburantes – Venda por Produtor ou Importador',
        '113': 'Álcool, Inclusive para Fins Carburantes – Venda por Distribuidor ou Comerciante Não Varejista',
        '114': 'Etanol Não Combustível – PJ não optante pelo regime especial em 2025',
        '115': 'Etanol Não Combustível – PJ optante pelo regime especial em 2025',
        '116': 'Etanol Não Combustível – A partir de 1º de janeiro de 2026',
        '117': 'Álcool, Inclusive para Fins Carburantes – Venda direta do produtor/importador (art. 68-B Lei 9.478/1997)',
        '150': 'Nafta Petroquímica e Condensado Destinados às Centrais Petroquímicas',
        '151': 'Etano, Propano, Butano e Correntes Gasosas de Refinaria – HLR – Destinados à Produção de Eteno, Propeno e outros',
        '152': 'Eteno, propeno, buteno, butadieno, orto-xileno, benzeno, tolueno, isopreno e paraxileno para indústrias químicas',
        '153': 'Vendas de gás natural e amônia para produção de cianeto de sódio, ácido cianídrico e outros',
        '199': 'Revenda de combustíveis – Alíquota zero (01/2011 a 01/2013)',
        '201': 'Produtos Farmacêuticos',
        '202': 'Produtos de Perfumaria, de Toucador ou de Higiene Pessoal',
        '299': 'Revenda de fármacos e perfumarias – Alíquota zero (01/2011 a 01/2013)',
        '301': 'Veículos Automotores e Máquinas Agrícolas',
        '302': 'Autopeças – Vendas para Atacadistas, Varejistas e Consumidores',
        '303': 'Autopeças – Vendas para Fabricantes de Veículos e Máquinas e de Autopeças',
        '304': 'Pneumáticos (Pneus Novos e Câmaras-de-Ar)',
        '399': 'Revenda de veículos, máquinas e autopeças – Alíquota zero (01/2011 a 01/2013)',
        '401': 'Águas Minerais Artificiais e Águas Gaseificadas Artificiais',
        '402': 'Águas Minerais Naturais, Incluídas as Naturalmente Gaseificadas',
        '403': 'Refrigerantes',
        '404': 'Preparações Compostas, não Alcoólicas, para Elaboração de Bebida Refrigerante',
        '405': 'Refrescos, Isotônicos e Energéticos',
        '406': 'Cervejas de Malte e Cervejas Sem Álcool',
        '407': 'Chope e Cervejas de Malte Quando Vendidas a Granel',
        '411': 'Preparações compostas bebida refrigerante – Vol. até 500 ml – Vendas a PJ varejista ou consumidor final',
        '412': 'Preparações compostas bebida refrigerante – Vol. acima de 500 ml – Vendas a PJ varejista ou consumidor final',
        '413': 'Preparações compostas bebida refrigerante – Vol. até 500 ml – Vendas às demais PJ',
        '414': 'Preparações compostas bebida refrigerante – Vol. acima de 500 ml – Vendas às demais PJ',
        '415': 'Águas – Vol. até 500 ml – Vendas a PJ varejista ou consumidor final',
        '416': 'Águas – Vol. acima de 500 ml – Vendas a PJ varejista ou consumidor final',
        '417': 'Águas – Vol. até 500 ml – Vendas às demais PJ',
        '418': 'Águas – Vol. acima de 500 ml – Vendas às demais PJ',
        '419': 'Águas adicionadas de açúcar ou aromatizadas – Vol. até 500 ml – Vendas a PJ varejista ou consumidor final',
        '420': 'Águas adicionadas de açúcar ou aromatizadas – Vol. acima de 500 ml – Vendas a PJ varejista ou consumidor final',
        '421': 'Águas adicionadas de açúcar ou aromatizadas – Vol. até 500 ml – Vendas às demais PJ',
        '422': 'Águas adicionadas de açúcar ou aromatizadas – Vol. acima de 500 ml – Vendas às demais PJ',
        '423': 'Cervejas de malte – Vol. até 400 ml – Vendas a PJ varejista ou consumidor final',
        '424': 'Cervejas de malte – Vol. acima de 400 ml – Vendas a PJ varejista ou consumidor final',
        '425': 'Cervejas de malte – Vol. até 400 ml – Vendas às demais PJ',
        '426': 'Cervejas de malte – Vol. acima de 400 ml – Vendas às demais PJ',
        '427': 'Cervejas/chopes especiais (até 5.000.000 L) – Vol. até 400 ml – Vendas às PJ em geral (não varejistas)',
        '428': 'Cervejas/chopes especiais (até 5.000.000 L) – Vol. acima de 400 ml – Vendas às PJ em geral (não varejistas)',
        '429': 'Cervejas/chopes especiais (5.000.001 a 10.000.000 L) – Vol. até 400 ml – Vendas às PJ em geral (não varejistas)',
        '430': 'Cervejas/chopes especiais (5.000.001 a 10.000.000 L) – Vol. acima de 400 ml – Vendas às PJ em geral (não varejistas)',
        '431': 'Cervejas/chopes especiais (até 5.000.000 L) – Vol. até 400 ml – Vendas a PJ varejistas ou consumidor final',
        '432': 'Cervejas/chopes especiais (até 5.000.000 L) – Vol. acima de 400 ml – Vendas a PJ varejistas ou consumidor final',
        '433': 'Cervejas/chopes especiais (5.000.001 a 10.000.000 L) – Vol. até 400 ml – Vendas a PJ varejistas ou consumidor final',
        '434': 'Cervejas/chopes especiais (5.000.001 a 10.000.000 L) – Vol. acima de 400 ml – Vendas a PJ varejistas ou consumidor final',
        '499': 'Revenda de bebidas frias – Alíquota zero (01/2011 a 01/2013)',
    },

    # ── CST 04 — Tabela 4.3.10 (v1.25, 30/03/2026) ──────────────────────
    '04': {
        '001':   'Revenda de combustíveis – Alíquota zero',
        '002':   'Revenda de fármacos e perfumarias – Alíquota zero',
        '003':   'Revenda de veículos, máquinas e autopeças – Alíquota zero',
        '004':   'Revenda de bebidas frias – Alíquota zero (até 30/04/2015)',
        '101': 'Gasolinas, Exceto Gasolina de Aviação',
        '102': 'Óleo Diesel',
        '103': 'Gás Liquefeito de Petróleo – GLP',
        '104': 'Querosene de Aviação',
        '105': 'Correntes Destinadas à Formulação de Gasolinas',
        '106': 'Correntes Destinadas à Formulação de Óleo Diesel',
        '107': 'Nafta Petroquímica Destinada à Formulação de Gasolina ou de Óleo Diesel',
        '108': 'Nafta Petroquímica Destinada à Formulação Exclusivamente de Óleo Diesel',
        '109': 'Biodiesel',
        '112': 'Álcool, Inclusive para Fins Carburantes – Venda por Produtor ou Importador',
        '113': 'Álcool, Inclusive para Fins Carburantes – Venda por Distribuidor ou Comerciante Não Varejista',
        '199': 'Revenda de combustíveis – Alíquota zero (01/2011 a 01/2013)',
        '201': 'Produtos Farmacêuticos',
        '202': 'Produtos de Perfumaria, de Toucador ou de Higiene Pessoal',
        '299': 'Revenda de fármacos e perfumarias – Alíquota zero (01/2011 a 01/2013)',
        '301': 'Veículos Automotores e Máquinas Agrícolas',
        '302': 'Autopeças – Vendas para Atacadistas, Varejistas e Consumidores',
        '303': 'Autopeças – Vendas para Fabricantes de Veículos e Máquinas e de Autopeças',
        '304': 'Pneumáticos (Pneus Novos e Câmaras-de-Ar)',
        '399': 'Revenda de veículos, máquinas e autopeças – Alíquota zero (01/2011 a 01/2013)',
        '401': 'Águas Minerais Artificiais e Águas Gaseificadas Artificiais',
        '402': 'Águas Minerais Naturais, Incluídas as Naturalmente Gaseificadas',
        '403': 'Refrigerantes',
        '404': 'Preparações Compostas, não Alcoólicas, para Elaboração de Bebida Refrigerante',
        '405': 'Refrescos, Isotônicos e Energéticos',
        '406': 'Cervejas de Malte e Cervejas Sem Álcool',
        '407': 'Chope e Cervejas de Malte Quando Vendidas a Granel',
        '499': 'Revenda de bebidas frias – Alíquota zero (01/2011 a 01/2013)',
        '918': 'Receita de venda de bebidas frias por PJ varejista – Alíquota zero (a partir de 01/05/2015)',
    },

    # ── CST 06 — Tabela 4.3.13 (v1.34, 16/04/2026) ──────────────────────
    '06': {
        '101': 'Adubos ou fertilizantes classificados no Capítulo 31, exceto os produtos de uso veterinário, da TIPI, e suas matérias-primas',
        '102': 'Defensivos agropecuários classificados na posição 38.08 da TIPI e suas matérias-primas',
        '103': 'Sementes e mudas destinadas à semeadura e plantio, e produtos de natureza biológica utilizados em sua produção',
        '104': 'Corretivo de solo de origem mineral classificado no Capítulo 25 da TIPI',
        '105': 'Legumes de vagem, secos, em grão, mesmo pelados ou partidos; arroz; farinhas e sêmolas',
        '106': 'Inoculantes agrícolas produzidos a partir de bactérias fixadoras de nitrogênio, classificados no código 3002.90.99 da TIPI',
        '107': 'Vacinas para medicina veterinária',
        '108': 'Farinha, grumos e sêmolas, grãos esmagados ou em flocos, de milho (códigos 1102.20, 1103.13 e 1104.19 da TIPI)',
        '109': 'Pintos de 1 (um) dia',
        '110': 'Leite fluido pasteurizado ou industrializado, leite em pó, leite fermentado, bebidas e compostos lácteos e fórmulas infantis',
        '111': 'Queijos tipo mozarela, minas, prato, queijo de coalho, ricota, requeijão, queijo provolone, queijo parmesão e queijo fresco não maturado',
        '112': 'Soro de leite fluido a ser empregado na industrialização de produtos destinados ao consumo humano',
        '113': 'Farinha de trigo',
        '114': 'Trigo',
        '115': 'Pré-misturas próprias para fabricação de pão comum e pão comum',
        '116': 'Produtos hortícolas e frutas',
        '117': 'Ovos',
        '118': 'Venda de semens e embriões',
        '119': 'Massas alimentícias classificadas na posição 19.02 da TIPI',
        '120': 'Queijo do reino',
        '121': 'Carnes bovina, suína, ovina, caprina e de aves e produtos de origem animal',
        '122': 'Peixes e outros produtos',
        '123': 'Café classificado nos códigos 09.01 e 2101.1 da TIPI',
        '124': 'Açúcar classificado nos códigos 1701.14.00 e 1701.99.00 da TIPI',
        '125': 'Óleo de soja classificado na posição 15.07 da TIPI e outros óleos vegetais classificados nas posições 15.08 a 15.14 da TIPI',
        '126': 'Manteiga classificada no código 0405.10.00 da TIPI',
        '127': 'Margarina classificada no código 1517.10.00',
        '128': 'Sabões de toucador classificados no código 3401.11.90 Ex 01 da TIPI',
        '129': 'Produtos para higiene bucal ou dentária classificados na posição 33.06 da TIPI',
        '130': 'Papel higiênico classificado no código 4818.10.00 da TIPI',
        '201': 'Aeronaves classificadas na posição 88.02 da TIPI',
        '202': 'Partes, peças, ferramentais, componentes, insumos e serviços para manutenção de aeronaves',
        '203': 'Álcool anidro adicionado à gasolina, por distribuidores',
        '204': 'Álcool, inclusive para fins carburantes, em operações realizadas em bolsa de mercadorias e futuros',
        '205': 'Carvão mineral destinado à geração de energia elétrica',
        '206': 'Biodiesel fabricado a partir de matérias-primas produzidas nas regiões norte, nordeste e no semi-árido, por agricultor familiar enquadrado no PRONAF',
        '207': 'Valores recebidos pelos concessionários pela intermediação ou entrega dos veículos (posições 87.03 e 87.04 da TIPI)',
        '208': 'Veículos novos montados sobre chassis para transporte escolar para educação básica na zona rural',
        '209': 'Embarcações novas para transporte escolar para educação básica na zona rural',
        '210': 'Materiais e equipamentos para construção, conservação, modernização, conversão ou reparo de embarcações registradas no REB',
        '211': 'Veículos e carros blindados de combate para uso das Forças Armadas ou órgãos de segurança pública brasileiros',
        '212': 'Gás natural canalizado destinado à produção de energia elétrica pelo Programa Prioritário de Termoeletricidade',
        '213': 'Serviços de transporte ferroviário em sistema de trens de alta velocidade (TAV)',
        '214': 'Receitas de serviços regulares de transporte coletivo municipal rodoviário, metroviário e ferroviário (até 13/11/2014)',
        '215': 'Receita de serviços regulares de transporte coletivo municipal rodoviário, metroviário, ferroviário e aquaviário (a partir de 14/11/2014)',
        '216': 'Álcool, Inclusive para Fins Carburantes – Venda por Distribuidor',
        '217': 'Óleo Diesel – As operações no mercado interno ou sobre importação',
        '218': 'Correntes Destinadas Exclusivamente à Formulação de Óleo Diesel',
        '219': 'GLP quando destinado ao uso doméstico e envasado em recipientes de até treze quilogramas',
        '220': 'Biodiesel – As operações no mercado interno ou sobre importação',
        '221': 'Querosene de Aviação – As operações no mercado interno ou sobre importação',
        '222': 'Gás Liquefeito de Petróleo – GLP derivado de petróleo e de gás natural',
        '223': 'Receitas da venda de produtos classificados no Ex 01 do código 8503.00.90 da TIPI, exceto pás eólicas',
        '224': 'Receita ou faturamento na venda ou sobre a importação de gás natural veicular – GNV',
        '225': 'As operações no mercado interno ou sobre importação que envolvam gasolina e suas correntes, exceto de aviação',
        '226': 'As operações no mercado interno e importação que envolvam etanol, inclusive para fins carburantes',
        '227': 'As receitas decorrentes da atividade de transporte aéreo regular de passageiros',
        '228': 'Programa Mover – As receitas e ganhos líquidos do Fundo Nacional de Desenvolvimento Industrial e Tecnológico (FNDIT)',
        '229': 'Biodiesel – importação (Decreto nº 12.923, de 7 de abril de 2026)',
        '301': 'Produtos classificados na posição 87.13 da NCM (cadeiras de rodas e outros veículos)',
        '302': 'Artigos e aparelhos ortopédicos ou para fraturas classificados no código 90.21.10 da NCM',
        '303': 'Artigos e aparelhos de próteses classificados no código 90.21.3 da NCM',
        '304': 'Almofadas antiescaras classificadas nos Capítulos 39, 40, 63 e 94 da NCM',
        '305': 'Bens relacionados em ato do Poder Executivo para aplicação nas Unidades Modulares de Saúde',
        '306': 'Produtos químicos classificados no Capítulo 29 da NCM',
        '307': 'Produtos químicos intermediários de síntese, classificados no Capítulo 29 da NCM',
        '308': 'Produtos destinados ao uso em hospitais, clínicas e consultórios médicos e odontológicos',
        '309': 'Produtos classificados nos códigos 8443.32.22, 8469.00.39 Ex 01, 8714.20.00, 9021.40.00, 9021.90.82 e 9021.90.92 da TIPI',
        '310': 'Calculadoras equipadas com sintetizador de voz classificadas no código 8470.10.00 Ex 01 da TIPI',
        '311': 'Teclados com colmeia classificados no código 8471.60.52 da TIPI',
        '312': 'Indicadores ou apontadores – mouses – com entrada para acionador classificados no código 8471.60.53 da TIPI',
        '313': 'Linhas braile classificadas no código 8471.60.90 Ex 01 da TIPI',
        '314': 'Digitalizadores de imagens – scanners – equipados com sintetizador de voz classificados no código 8471.90.14 Ex 01 da TIPI',
        '315': 'Duplicadores braile classificados no código 8472.10.00 Ex 01 da TIPI',
        '316': 'Acionadores de pressão classificados no código 8471.60.53 Ex.02 da TIPI',
        '317': 'Lupas eletrônicas do tipo utilizado por pessoas com deficiência visual classificadas no código 8525.80.19 Ex 01 da TIPI',
        '318': 'Implantes cocleares classificados no código 9021.40.00 da TIPI',
        '319': 'Próteses oculares classificadas no código 9021.39.80 da TIPI',
        '320': 'Programas – softwares – de leitores de tela que convertem texto em voz sintetizada para auxílio de pessoas com deficiência visual',
        '321': 'Aparelhos contendo programas – softwares – de leitores de tela que convertem texto em caracteres braile, para utilização de surdos-cegos',
        '322': 'Neuroestimuladores para tremor essencial/Parkinson, classificados no código 9021.90.19, e seus acessórios',
        '323': 'Equipamentos ou materiais destinados a uso médico, hospitalar, clínico ou laboratorial, quando adquiridos pela União, Estados, DF ou Municípios, ou por entidades beneficentes',
        '324': 'Venda no mercado interno ou a importação de sulfato de zinco para medicamentos utilizados em nutrição parenteral',
        '401': 'Venda a varejo de unidades de processamento digital classificadas no código 8471.50.10 da TIPI (até 30/11/2015)',
        '402': 'Venda a varejo de máquinas automáticas para processamento de dados, digitais, portáteis (até 30/11/2015)',
        '403': 'Venda a varejo de máquinas automáticas de processamento de dados apresentadas sob a forma de sistemas (até 30/11/2015)',
        '404': 'Venda a varejo de teclado e mouse quando acompanharem a unidade de processamento digital (até 30/11/2015)',
        '405': 'PADIS – Programa de Apoio ao Desenvolvimento Tecnológico da Indústria de Semicondutores',
        '406': 'PATVD – Programa de Apoio ao Desenvolvimento Tecnológico da Indústria de Equipamentos para a TV Digital',
        '407': 'Venda a varejo de Tablet PC (até 30/11/2015)',
        '408': 'Venda a varejo de modems (até 30/11/2015)',
        '409': 'Venda a varejo de smartphones (até 30/11/2015)',
        '410': 'Venda a varejo de roteadores digitais (até 30/11/2015)',
        '411': 'Venda dos produtos relacionados nos códigos 401 a 410 a PJ de direito privado ou órgãos e entidades da Administração Pública (até 30/11/2015)',
        '412': 'Venda dos produtos relacionados nos códigos 401 a 410 a sociedades de arrendamento mercantil leasing (até 30/11/2015)',
        '413': 'Receita decorrente da venda de bens de defesa nacional por PJ beneficiária do RETID à União',
        '414': 'Receita decorrente da prestação de serviços de tecnologia industrial básica por PJ beneficiária do RETID à União',
        '901': 'Papel destinado à impressão de jornais (até 30/04/2016)',
        '902': 'Papéis classificados nos códigos 4801.00.10, 4801.00.90, 4802.61.91, 4802.61.99, 4810.19.89 e 4810.22.90 da TIPI, destinados à impressão de periódicos (até 30/04/2016)',
        '903': 'Livros, conforme definido no art. 2º da Lei nº 10.753/03',
        '904': 'Preparações compostas não-alcoólicas, classificadas no código 2106.90.10 Ex 01 da TIPI, destinadas à elaboração de bebidas',
        '905': 'Material de defesa, classificado nas posições 87.10.00.00 e 89.06.10.00 da TIPI',
        '906': 'Equipamentos de controle de produção, inclusive medidores de vazão, condutivímetros',
        '907': 'Valores pagos ou creditados pelos Estados, DF e Municípios relativos ao ICMS e ao ISS, no âmbito de programas de concessão de crédito',
        '908': 'Vendas de mercadorias destinadas ao consumo ou à industrialização na Zona Franca de Manaus – ZFM',
        '909': 'Vendas de mercadorias destinadas ao consumo ou à industrialização nas Áreas de Livre Comércio – ALC',
        '910': 'Vendas de matérias-primas, produtos intermediários e materiais de embalagem produzidos na ZFM para estabelecimentos industriais ali instalados',
        '911': 'Receitas financeiras, inclusive decorrentes de operações realizadas para fins de hedge, auferidas pelas pessoas jurídicas sujeitas ao regime não-cumulativo',
        '912': 'Aquisição no mercado interno ou a importação de mercadoria equivalente à empregada ou consumida na industrialização de produto exportado (Drawback Reposição de Estoque)',
        '913': 'Projetores para exibição cinematográfica, classificados no código 9007.2 da NCM, e suas partes e acessórios',
        '914': 'Receita decorrente da venda de águas minerais naturais comercializadas em recipientes com capacidade nominal inferior a 10 litros ou igual ou superior a 10 litros',
        '915': 'Concessões de geração, transmissão e distribuição de energia elétrica – Valor da indenização (art. 8º, § 4º, da Lei nº 12.783/2013)',
        '916': 'Concessões de geração, transmissão e distribuição de energia elétrica – Valor da indenização (art. 15, § 9º, da Lei nº 12.783/2013)',
        '917': 'Valores efetivamente recebidos exclusivamente a título da subvenção extraordinária aos produtores fornecedores independentes de cana-de-açúcar',
        '918': 'Receita decorrente da venda de bebidas frias por pessoa jurídica varejista (a partir de 01/05/2015)',
        '919': 'Receita auferida na venda de pneumáticos e câmaras de ar de borracha para bicicletas classificados nos códigos 4011.50.00 e 4013.20.00 da TIPI',
        '920': 'Perse – Programa Emergencial de Retomada do Setor de Eventos (03/2022 a 02/2027)',
        '999': 'Código genérico – Operações tributáveis a alíquota zero',
    },

    # ── CST 07 — Tabela 4.3.14 (28/04/2026) ─────────────────────────────
    '07': {
        '101': 'Fornecimento de mercadorias ou serviços para uso ou consumo de bordo em embarcações e aeronaves em tráfego internacional (exceto querosene de aviação)',
        '102': 'Transporte internacional de cargas ou passageiros',
        '103': 'Receitas auferidas pelos estaleiros navais brasileiros nas atividades de construção, conservação, modernização, conversão e reparo de embarcações (REB)',
        '104': 'Frete de mercadorias transportadas entre o País e o exterior pelas embarcações registradas no REB',
        '201': 'Receitas relativas às atividades próprias dos templos de qualquer culto; partidos políticos; instituições de educação e de assistência social; sindicatos; fundações de direito privado; condomínios; OCB e Organizações Estaduais de Cooperativas',
        '202': 'Receitas das entidades beneficentes de assistência social com a finalidade de prestação de serviços nas áreas de assistência social, saúde ou educação (art. 29 da Lei nº 12.101/2009)',
        '301': 'Venda de energia elétrica pela Itaipu Binacional',
        '401': 'Importação de bens ou mercadorias para uso exclusivo na organização e realização dos eventos FIFA (Copa das Confederações 2013, Copa do Mundo 2014)',
        '402': 'Receita auferida por Subsidiária FIFA no Brasil, decorrente das atividades próprias e diretamente vinculadas à organização ou realização dos referidos eventos',
        '403': 'Receita das atividades próprias, auferida pelos Prestadores de Serviços da FIFA, estabelecidos no País',
        '404': 'Importações de bens, mercadorias ou serviços para uso exclusivo em atividades próprias e diretamente vinculadas à organização ou realização dos Eventos Olímpicos de 2016',
        '405': 'Rendimentos, remessas e operações de câmbio e seguros, decorrentes das atividades próprias e diretamente vinculadas à organização ou realização dos eventos olímpicos de 2016, envolvendo o CIO',
        '406': 'Receitas e rendimentos auferidos pelas empresas vinculadas ao CIO, domiciliadas no Brasil, em relação aos fatos geradores decorrentes das atividades dos Eventos Olímpicos de 2016',
        '407': 'Receitas, lucros e rendimentos auferidos pelo RIO 2016, em relação aos fatos geradores decorrentes das atividades próprias e diretamente vinculadas à organização ou realização do evento',
        '901': 'Recursos recebidos a título de repasse, oriundos do Orçamento Geral da União, dos Estados, do Distrito Federal e dos Municípios, pelas empresas públicas e sociedades de economia mista',
        '902': 'Receita da instituição privada de ensino superior, com fins lucrativos ou sem fins lucrativos não beneficente, que aderir ao Programa Universidade para Todos (Prouni)',
        '903': 'Receita bruta de venda a varejo dos componentes e equipamentos de rede, terminais e transceptores dedicados aos serviços de telecomunicações (subfaixas de radiofrequência de 451 MHz a 468 MHz)',
        '904': 'Desperdícios, resíduos ou aparas de plástico, de papel ou cartão, de vidro, de ferro ou aço, de cobre, de níquel, de alumínio e outros metálicos (Lei nº 11.196/2005, art. 48, na redação dada pela Lei nº 15.394, de 22/04/2026)',
        '999': 'Código genérico – Operações com Isenção',
    },

    # ── CST 08 — Tabela 4.3.15 (v1.02) ──────────────────────────────────
    '08': {
        '101': 'Venda de querosene de aviação por pessoa jurídica não enquadrada na condição de importadora ou produtora',
        '102': 'Venda de querosene de aviação por produtora ou importadora a distribuidora, quando o produto for destinado ao consumo por aeronave em tráfego internacional',
        '201': 'Vendas de biodiesel por pessoas não enquadradas como produtor ou importador',
        '301': 'Vendas de materiais e equipamentos, bem assim da prestação de serviços decorrentes dessas operações, efetuadas diretamente a Itaipu Binacional',
        '401': 'Exportação de mercadorias para o exterior',
        '402': 'Serviços prestados a pessoas físicas ou jurídicas residentes ou domiciliadas no exterior, cujo pagamento represente ingresso de divisas',
        '403': 'Vendas, com o fim específico de exportação, a empresa comercial exportadora constituída nos termos do Decreto-Lei nº 1.248/1972, ou simplesmente registrada na Secex',
        '901': 'Regime Cumulativo – Demais receitas não classificadas como faturamento, não enquadradas como receita bruta nos termos do art. 3º da Lei nº 9.718/1998',
        '902': 'Receitas de venda de crédito de carbono, dos ativos CBE e CRVEs (a partir de 12/2024)',
        '999': 'Outras receitas sem incidência',
    },

    # ── CST 09 — Tabela 4.3.16 (28/04/2026) ─────────────────────────────
    '09': {
        '101': 'Vendas a pessoa jurídica preponderantemente exportadora',
        '102': 'Vendas a fabricante de veículos e carros blindados de combate (NCM 8710.00.00) para uso pelas Forças Armadas ou órgãos de segurança pública brasileiros',
        '103': 'Aquisição no mercado interno ou a importação de mercadoria para emprego ou consumo na industrialização de produto a ser exportado (Drawback Integrado)',
        '104': 'Aquisição no mercado interno ou à importação de mercadorias para emprego em reparo, criação, cultivo ou atividade extrativista de produto a ser exportado',
        '105': 'Aquisição no mercado interno ou importações de empresas denominadas fabricantes-intermediários, para industrialização de produto intermediário (Drawback Intermediário)',
        '106': 'A aquisição no mercado interno ou a importação de serviço direta e exclusivamente vinculado à exportação ou entrega no exterior de produto resultante da utilização do regime (art. 12 da Lei 11.945/2009, a partir de 01/2023)',
        '201': 'Insumos de origem animal, utilizados na fabricação de produtos destinados à alimentação humana ou animal (01/2011 a 12/2011)',
        '202': 'Insumos de origem vegetal, utilizados na fabricação de produtos destinados à alimentação humana ou animal',
        '203': 'Soja e seus derivados classificados nos Capítulos 12, 15 e 23, todos da TIPI (01/2011 a 09/10/2013)',
        '204': 'Venda de cerealista que exerça cumulativamente as atividades de limpar, padronizar, armazenar e comercializar os produtos in natura de origem vegetal',
        '205': 'Venda a granel de leite in natura, efetuada por pessoa jurídica que exerça cumulativamente as atividades de transporte e resfriamento deste produto',
        '206': 'Venda por PJ que exerça atividade agropecuária ou por cooperativa de produção agropecuária de produto in natura de origem vegetal destinado à elaboração de mercadorias classificadas no código 22.04 da NCM',
        '207': 'Venda de animais vivos classificados nas posições 01.02 e 01.04 à pessoa jurídica que produza mercadorias classificadas nos códigos 02.01, 02.02, 02.04 e outros da NCM',
        '208': 'Vendas de produtos classificados nas posições 02.01, 02.02, 0206.10.00, 0206.20, 0206.21, 0206.29, 0210.20.00, 0506.90.00, 0510.00.10 e 1502.00.1 (até 01/09/2013)',
        '209': 'Receita bruta da venda, no mercado interno, de insumos de origem vegetal para alimentação de suínos e aves, preparações e animais vivos e produtos derivados de suínos e aves',
        '210': 'Receitas decorrentes da venda dos produtos classificados nos códigos 0901.1 e 0901.90.00 da TIPI, exceto na venda a consumidor final (01/2012 a 08/03/2013)',
        '211': 'Receitas decorrentes da venda de matéria-prima in natura de origem vegetal, destinada à produção de biodiesel',
        '212': 'Receitas decorrentes da venda dos produtos classificados no código 0805.10.00 da TIPI, quando utilizados na industrialização dos produtos classificados no código 2009.1 da TIPI, destinados à exportação (a partir de 01/2013)',
        '213': 'Soja classificada na posição 12.01 e dos produtos classificados nos códigos 1208.10.00 e 2304.00 da TIPI (a partir de 10/10/2013)',
        '301': 'REPES – Regime Especial de Tributação para a Plataforma de Exportação de Serviços de Tecnologia da Informação',
        '302': 'RECAP – Regime Especial de Aquisição de Bens de Capital para Empresas Exportadoras',
        '303': 'REIDI – Regime Especial de Incentivos para o Desenvolvimento da Infra-Estrutura',
        '304': 'REPENEC – Regime Especial de Incentivos para o Desenvolvimento de Infraestrutura da Indústria Petrolífera nas Regiões Norte, Nordeste e Centro-Oeste',
        '305': 'REPORTO – Regime Tributário para Incentivo à Modernização e à Ampliação da Estrutura Portuária',
        '306': 'RECOMPE – Regime Especial de Aquisição de Computadores para Uso Educacional',
        '307': 'RETAERO – Regime Especial para a Indústria Aeronáutica Brasileira',
        '308': 'RECOPA – Regime Especial de Tributação para Construção, Ampliação, Reforma ou Modernização de Estádios de Futebol',
        '309': 'ZFM – Zona Franca de Manaus – Importação de bens a serem empregados na elaboração de matérias-primas destinadas a emprego em processo de industrialização por estabelecimentos industriais instalados na ZFM',
        '310': 'ZPE – Zonas de Processamento de Exportação – Importações ou aquisições no mercado interno de bens e serviços por empresa autorizada a operar em ZPE',
        '311': 'Vendas realizadas no mercado interno para a FIFA, para Subsidiária FIFA no Brasil ou para a Emissora Fonte da FIFA, de mercadorias destinadas a uso ou consumo exclusivo na organização e realização da Copa das Confederações FIFA 2013 e da Copa do Mundo FIFA 2014',
        '312': 'RECOF – Regime de Entreposto Industrial sob Controle Aduaneiro Informatizado',
        '313': 'RECOM – Regime Aduaneiro Especial de Importação de Insumos Destinados a Industrialização por Encomenda de Produtos Classificados nas Posições 8701 A 8705 da NCM',
        '314': 'RECINE – Regime Especial de Tributação para Desenvolvimento da Atividade de Exibição Cinematográfica',
        '315': 'PROUCA/REICOMP – Programa Um Computador por Aluno / Regime Especial de Incentivo a Computadores para Uso Educacional (até 31/12/2015)',
        '316': 'REPNBL-Redes – Regime Especial de Tributação do Programa Nacional de Banda Larga para Implantação de Redes de Telecomunicações (até 31/12/2016)',
        '317': 'RETID – Regime Especial Tributário para a Indústria de Defesa',
        '318': 'REIF – Regime Especial de Incentivo ao Desenvolvimento da Infraestrutura da Indústria de Fertilizantes',
        '319': 'Vendas de mercadorias e a prestação de serviços ocorridas no mercado interno para as pessoas jurídicas mencionadas no § 2º do art. 4º da Lei nº 12.780/2013, destinadas exclusivamente à organização ou à realização dos Jogos Olímpicos de 2016 e dos Jogos Paraolímpicos de 2016 (até 31/12/2017)',
        '320': 'REPETRO-Industrialização – Venda no mercado interno de matérias-primas, produtos intermediários e materiais de embalagem para serem utilizados integralmente no processo de industrialização de produto final destinado às atividades de exploração, de desenvolvimento e de produção de petróleo, de gás natural e de outros hidrocarbonetos fluidos',
        '321': 'REPETRO-SPED – Venda dos produtos finais destinados às atividades de exploração, de desenvolvimento e de produção de petróleo, de gás natural e de outros hidrocarbonetos fluidos por fabricantes desses, beneficiários do Repetro-Industrialização',
        '401': 'Receitas de Fretes e de transporte multimodal, contratadas por pessoa jurídica preponderantemente exportadora, para transporte no mercado interno de produtos com suspensão ou destinados a Exportação',
        '402': 'Venda de cana-de-açúcar, classificada na posição 12.12 da NCM, efetuada para pessoa jurídica produtora de álcool, inclusive para fins carburantes, tributada no regime de não cumulatividade',
        '403': 'Venda de óleo combustível, tipo bunker, MF – Marine Fuel, MGO – Marine Gás Oil e ODM – Óleo Diesel Marítimo, quando destinados à navegação de cabotagem e de apoio portuário e marítimo, para a pessoa jurídica previamente habilitada',
        '404': 'Acetona classificada no código 2914.11.00 da TIPI, destinada à produção de monoisopropilamina (Mipa) utilizada na elaboração de defensivos agropecuários classificados na posição 38.08 da TIPI',
        '405': 'Desperdícios, resíduos ou aparas de plástico, de papel ou cartão, de vidro, de ferro ou aço, de cobre, de níquel, de alumínio, de chumbo, de zinco e de estanho, e demais desperdícios e resíduos metálicos do Capítulo 81 da TIPI, quando vendidos para pessoa jurídica que apure o imposto de renda com base no lucro real (até 22/04/2026)',
        '406': 'Venda de produtos à pessoa jurídica sediada no exterior, com contrato de entrega no território nacional, de insumos destinados à industrialização, por conta e ordem da encomendante sediada no exterior, de máquinas e veículos classificados nas posições 87.01 a 87.05 da TIPI',
        '407': 'Vendas a empresa sediada no exterior, para entrega em território nacional, de material de embalagem a ser totalmente utilizados no acondicionamento de mercadoria destinada à exportação para o exterior',
        '408': 'Venda de máquinas e equipamentos classificados na posição 84.39, utilizados na fabricação de papéis destinados à impressão de jornais ou de papéis destinados à impressão de periódicos',
        '409': 'As aquisições no mercado interno e nas importações de petróleo efetuadas por refinarias para a produção de combustíveis',
        '410': 'Aquisições dos insumos naftas NCM/SH 2710.12.49, outras misturas (aromáticos) NCM/SH 2707.99.90, óleo de petróleo parcialmente refinado NCM 2710.19.99, outros óleos brutos de petróleo ou minerais (condensados) NCM 2709.00.10, e N-Metilanilina NCM/SH 2921.42.90',
        '901': 'Doações em espécie recebidas por instituições financeiras públicas controladas pela União e destinadas a ações de prevenção, monitoramento e combate ao desmatamento',
        '999': 'Outras operações com suspensão',
    },
}

CST_DESCRICOES = {
    '01': 'CST 01 – Tributável Alíquota Básica',
    '02': 'CST 02 – Tributável Alíquota Diferenciada',
    '03': 'CST 03 – Tributável por Unidade de Medida',
    '04': 'CST 04 – Monofásica / Revenda Alíquota Zero',
    '05': 'CST 05 – Substituição Tributária',
    '06': 'CST 06 – Alíquota Zero',
    '07': 'CST 07 – Isenta da Contribuição',
    '08': 'CST 08 – Sem Incidência',
    '09': 'CST 09 – Suspensão da Contribuição',
    '49': 'CST 49 – Outras Operações de Saída',
    '50': 'CST 50 – Aquisição com Direito a Crédito',
    '70': 'CST 70 – Aquisição sem Direito a Crédito',
    '98': 'CST 98 – Outras Operações de Entrada',
    '99': 'CST 99 – Outras Operações',
}

BASE_CREDITO_OPCOES = {
    '':   '-- Não informar --',
    '01': '01 – Aquisição de bens para revenda',
    '02': '02 – Aquisição de bens utilizados como insumo',
    '03': '03 – Aquisição de serviços utilizados como insumo',
    '04': '04 – Energia elétrica e térmica',
    '05': '05 – Aluguéis de prédios',
    '06': '06 – Aluguéis de máquinas e equipamentos',
    '07': '07 – Armazenagem de mercadoria e frete na operação de venda',
    '08': '08 – Contraprestações de arrendamento mercantil',
    '09': '09 – Máquinas e equipamentos (crédito sobre depreciação)',
    '10': '10 – Máquinas e equipamentos (crédito sobre valor de aquisição)',
    '11': '11 – Amortização de edificações e benfeitorias em imóveis',
    '12': '12 – Devolução de vendas sujeitas à incidência não-cumulativa',
    '13': '13 – Outras operações com direito a crédito',
    '14': '14 – Atividade de transporte de cargas – subcontratação',
    '15': '15 – Atividade imobiliária – custo incorrido',
    '16': '16 – Atividade imobiliária – custo orçado',
    '17': '17 – Serviços de limpeza, conservação e manutenção',
    '18': '18 – Estoque de abertura de bens',
}

VINCULO_CREDITO_OPCOES = {
    '':   '-- Não informar --',
    '01': '01 – Crédito vinculado à alíquota básica',
    '02': '02 – Crédito vinculado à alíquota diferenciada',
    '03': '03 – Crédito vinculado à alíquota por unidade de produto',
    '05': '05 – Crédito vinculado a aquisição de embalagem',
    '06': '06 – Crédito presumido agroindústria e aquisição de combustível',
    '08': '08 – Crédito de importação',
    '99': '99 – Outros créditos',
}

# ==============================
# HELPERS
# ==============================
def _aliq_float(valor: str) -> float:
    v = str(valor).strip().replace(',', '.')
    if not v or v.upper() == 'NAN': return 0.0
    try:
        f = float(v)
        if 0 < f < 0.20: f = round(f * 100, 6)
        return f
    except ValueError: return 0.0


def get_codigo_pis(aliq_str: str, por_nome: dict) -> int:
    aliq = _aliq_float(aliq_str)
    for (central, tol, cod) in ALIQ_PIS_COFINS['PIS']:
        if abs(aliq - central) <= tol: return cod
    return get_codigo_imposto('PIS', por_nome, 4)


def get_codigo_cofins(aliq_str: str, por_nome: dict) -> int:
    aliq = _aliq_float(aliq_str)
    for (central, tol, cod) in ALIQ_PIS_COFINS['COFINS']:
        if abs(aliq - central) <= tol: return cod
    return get_codigo_imposto('COFINS', por_nome, 5)


def converter_data(data_sped: str) -> str:
    """DDMMAAAA ou DD/MM/AAAA → DD/MM/AAAA"""
    d = str(data_sped).strip().replace('/', '').replace('-', '')
    if len(d) == 8 and d.isdigit():
        return f"{d[0:2]}/{d[2:4]}/{d[4:8]}"
    return data_sped


def converter_data_planilha(valor) -> str:
    """Pandas Timestamp / ISO string → DD/MM/AAAA"""
    try:
        if pd.isna(valor): return ''
    except Exception: pass
    s = str(valor).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d%m%Y'):
        try: return datetime.strptime(s, fmt).strftime('%d/%m/%Y')
        except ValueError: continue
    return s


def _normalizar_cst(cst: str) -> str:
    s = str(cst).strip().split('.')[0]
    try: return str(int(s)).zfill(2)
    except ValueError: return s.zfill(2)


def _cst_saida_para_entrada(cst_saida: str) -> str:
    """
    Converte CST de saída para o CST de entrada equivalente.
    CSTs de entrada (50-99) ficam como estão.
    CSTs de saída (01-09, 49, 99) ficam como estão.
    """
    c = _normalizar_cst(cst_saida)
    try:
        n = int(c)
    except ValueError:
        return c
    # CSTs de entrada já são 50-99
    if n >= 50:
        return c
    # CSTs de saída: retorna o mesmo (já são CSTs de saída)
    return c


def _cst_entrada_para_saida(cst_entrada: str) -> str:
    """
    Converte CST de entrada (50-99) para o CST de saída equivalente.
    Usado no cadastro do produto (0110 campo 16 = CST Saídas).
    """
    c = _normalizar_cst(cst_entrada)
    return CST_ENTRADA_PARA_SAIDA.get(c, c)

 ==============================

# DE-PARA INVERSO: CST SAÍDA → CST ENTRADA
# Conforme arquivo "CST DE-Para entrada e saida.xlsx"
# Usado no campo 3 do 0110 (CST Entrada) quando o produto
# possui operação de saída com o CST correspondente.
# ==============================
CST_SAIDA_PARA_ENTRADA = {
    '01': '50',  # Saída tributada alíq. básica       → Aquisição com direito a crédito básica
    '02': '51',  # Saída tributada alíq. diferenciada → Aquisição com direito a crédito diferenciada
    '08': '74',  # Saída sem incidência               → Aquisição sem incidência
    '06': '73',  # Saída alíquota zero                → Aquisição com alíquota zero
    '04': '70',  # Saída monofásica                   → Aquisição monofásica/substituição tributária
    '49': '99',  # Outras saídas não tributadas        → Outras aquisições sem crédito
}


def _cst_saida_para_entrada_depara(cst_saida: str) -> str:
    """
    Converte CST de saída (01-49) para o CST de entrada equivalente
    conforme De-Para oficial. Usado no campo 3 do 0110 (CST Entrada)
    quando o produto é identificado por uma nota de saída com CST 01.
    Se não houver mapeamento, retorna o CST de saída sem alteração.
    """
    c = _normalizar_cst(cst_saida)
    return CST_SAIDA_PARA_ENTRADA.get(c, c)


def get_natureza_por_cst(cst_pis: str, cst_cofins: str, config_natureza: dict) -> tuple:
    cst_p = _normalizar_cst(cst_pis)
    cst_c = _normalizar_cst(cst_cofins)
    # Para saídas, a natureza é buscada pelo CST de saída
    # Para entradas, converte para saída antes de buscar
    cst_p_saida = _cst_entrada_para_saida(cst_p) if int(cst_p) >= 50 else cst_p
    cst_c_saida = _cst_entrada_para_saida(cst_c) if int(cst_c) >= 50 else cst_c
    nat_pis     = config_natureza.get(f"cst_{cst_p_saida}_nat_pis",    '')
    nat_cofins  = config_natureza.get(f"cst_{cst_c_saida}_nat_cofins", '')
    base_pis    = config_natureza.get(f"cst_{cst_p}_base_pis",         '')
    base_cofins = config_natureza.get(f"cst_{cst_c}_base_cofins",      '')
    vinc_pis    = config_natureza.get(f"cst_{cst_p}_vinc_pis",         '')
    vinc_cofins = config_natureza.get(f"cst_{cst_c}_vinc_cofins",      '')
    return nat_pis, nat_cofins, base_pis, base_cofins, vinc_pis, vinc_cofins


# ==============================
# TEMA TR
# ==============================
def apply_tr_theme():
    st.markdown("""
        <style>
        html, body, [class*="css"] {
            font-family: 'Segoe UI', 'Arial', sans-serif; color: #444444;
        }
        h1, h2, h3 { color: #FF8000; font-weight: 700; }
        section[data-testid="stSidebar"] { background-color: #444444; color: #FFFFFF; }
        section[data-testid="stSidebar"] * { color: #FFFFFF !important; }
        .stButton > button {
            background-color: #FF8000; color: #FFFFFF;
            border: none; border-radius: 4px; font-weight: bold;
        }
        .stButton > button:hover { background-color: #D64001; }
        .stDownloadButton > button {
            background-color: #FF8000; color: #FFFFFF;
            border: none; border-radius: 4px; font-weight: bold;
        }
        .stDownloadButton > button:hover { background-color: #D64001; }
        hr { border-color: #FF8000; }
        [data-testid="metric-container"] {
            background-color: #E9E9E9; border-left: 4px solid #FF8000;
            border-radius: 4px; padding: 10px;
        }
        .natureza-box {
            background-color: #FFF8F0; border-left: 4px solid #FF8000;
            border-radius: 4px; padding: 14px 18px; margin: 8px 0;
        }
        .cst-header-obrig {
            background: linear-gradient(90deg, #B71C1C 0%, #D32F2F 100%);
            color: white; padding: 8px 14px; border-radius: 4px;
            font-weight: bold; font-size: 13px; margin: 8px 0 4px 0;
        }
        .natureza-ativa {
            background-color: #E8F5E9; border-left: 3px solid #388E3C;
            padding: 5px 10px; border-radius: 4px; margin: 2px 0; font-size: 12px;
        }
        .cbenef-box {
            background-color: #E3F2FD; border-left: 4px solid #1565C0;
            border-radius: 4px; padding: 12px 16px; margin: 8px 0; font-size: 13px;
        }
        .info-ok-box {
            background-color: #E8F5E9; border-left: 4px solid #2E7D32;
            border-radius: 4px; padding: 10px 14px; margin: 6px 0; font-size: 13px;
        }
        .upload-row {
            background-color: #F5F5F5; border-radius: 8px;
            padding: 16px; margin: 8px 0;
        }
        </style>
    """, unsafe_allow_html=True)


def decode_arquivo(raw: bytes) -> str:
    for enc in ('utf-8', 'latin-1', 'cp1252'):
        try: return raw.decode(enc)
        except UnicodeDecodeError: continue
    return raw.decode('utf-8', errors='replace')


def encode_ansi_seguro(conteudo: str, log: list) -> bytes:
    resultado = []; substituicoes = 0
    for char in conteudo:
        try: resultado.append(char.encode('latin-1'))
        except UnicodeEncodeError: resultado.append(b'?'); substituicoes += 1
    if substituicoes:
        log.append(f"AVISO: {substituicoes} caractere(s) fora do ANSI substituídos por '?'.")
    return b''.join(resultado)


def _candidatos(nome: str) -> list:
    try: base = os.path.dirname(os.path.abspath(__file__))
    except NameError: base = os.getcwd()
    return [os.path.join(base, nome), os.path.join(os.getcwd(), nome), nome]


@st.cache_data(show_spinner=False)
def carregar_tabela_impostos() -> tuple:
    caminho = next((c for c in _candidatos(NOME_IMP_XLSX) if os.path.isfile(c)), None)
    por_codigo: dict = {}; por_nome: dict = {}
    if caminho is None:
        for nome, cod in IMPOSTOS_FALLBACK.items():
            por_codigo[cod] = nome; por_nome[nome.upper()] = cod
        return por_codigo, por_nome
    try:
        df = pd.read_excel(caminho, dtype=str)
        df.columns = [str(c).strip().upper() for c in df.columns]
        col_cod  = next((c for c in df.columns if 'CÓD' in c or 'COD' in c or c == 'CÓDIGO'), None)
        col_nome = next((c for c in df.columns if 'NOME' in c), None)
        if col_cod is None or col_nome is None:
            if len(df.columns) >= 2: col_cod, col_nome = df.columns[0], df.columns[1]
            else:
                for nome, cod in IMPOSTOS_FALLBACK.items():
                    por_codigo[cod] = nome; por_nome[nome.upper()] = cod
                return por_codigo, por_nome
        for _, row in df.iterrows():
            raw_cod = str(row[col_cod]).strip().split('.')[0]
            nome    = str(row[col_nome]).strip()
            if not raw_cod or raw_cod.upper() == 'NAN' or not nome or nome.upper() == 'NAN': continue
            try: cod = int(raw_cod)
            except ValueError: continue
            por_codigo[cod] = nome; por_nome[nome.upper()] = cod
    except Exception:
        for nome, cod in IMPOSTOS_FALLBACK.items():
            por_codigo[cod] = nome; por_nome[nome.upper()] = cod
    return por_codigo, por_nome


def get_codigo_imposto(nome_fragmento: str, por_nome: dict, default: int = 0) -> int:
    chave = nome_fragmento.strip().upper()
    if chave in por_nome: return por_nome[chave]
    for k, v in por_nome.items():
        if chave in k or k in chave: return v
    return default


@st.cache_data(show_spinner=False)
def carregar_tabela_cfop_oficial() -> tuple:
    caminho = next((c for c in _candidatos(NOME_CFOP_XLSX) if os.path.isfile(c)), None)
    tabela_descr: dict = {}; tabela_flags: dict = {}
    if caminho is None: return tabela_descr, tabela_flags
    try:
        df = pd.read_excel(caminho, sheet_name="CFOP", dtype=str)
        df.columns = [str(c).strip().upper() for c in df.columns]
        col_cfop  = next((c for c in df.columns if c == 'CFOP'), None)
        col_descr = next((c for c in df.columns if 'DESCRI' in c or 'RESUMIDA' in c), None)
        col_nfe   = next((c for c in df.columns if 'INDNFE'      in c.replace(' ', '').upper()), None)
        col_com   = next((c for c in df.columns if 'INDCOMUNICA' in c.replace(' ', '').upper()), None)
        col_trp   = next((c for c in df.columns if 'INDTRANSP'   in c.replace(' ', '').upper()), None)
        col_dev   = next((c for c in df.columns if 'INDDEVOL'    in c.replace(' ', '').upper()), None)
        if col_cfop is None or col_descr is None: return tabela_descr, tabela_flags
        def _flag(row, col):
            if col is None: return 0
            v = str(row[col]).strip().split('.')[0]
            try: return int(v)
            except ValueError: return 0
        for _, row in df.iterrows():
            raw = str(row[col_cfop]).strip()
            if '.' in raw: raw = raw.split('.')[0]
            raw = ''.join(filter(str.isdigit, raw))
            if not raw: continue
            cfop  = raw.zfill(4)
            descr = str(row[col_descr]).strip()
            if cfop == '0000' or not descr or descr.lower() == 'nan': continue
            tabela_descr[cfop] = descr
            tabela_flags[cfop] = {
                'indNFe': _flag(row, col_nfe), 'indComunica': _flag(row, col_com),
                'indTransp': _flag(row, col_trp), 'indDevol': _flag(row, col_dev),
            }
    except Exception: return {}, {}
    return tabela_descr, tabela_flags


def get_descricao_cfop(cfop: str, tabela_descr: dict) -> str:
    return tabela_descr.get(str(cfop).strip().zfill(4), '— descrição não encontrada —')

def get_flags_cfop(cfop: str, tabela_flags: dict) -> dict:
    return tabela_flags.get(str(cfop).strip().zfill(4),
                            {'indNFe': 0, 'indComunica': 0, 'indTransp': 0, 'indDevol': 0})

def get_tipo_operacao(cfop: str) -> str:
    p = str(cfop).strip()[:1]
    if p in ('1', '2', '3'): return 'Entrada'
    if p in ('5', '6', '7'): return 'Saída'
    return 'Desconhecido'

def is_devolucao(cfop: str, tabela_flags: dict) -> bool:
    return get_flags_cfop(cfop, tabela_flags).get('indDevol', 0) == 1


# ==============================
# CARREGAMENTO DA PLANILHA CLIENTE
# ==============================
def carregar_planilha_cliente(arquivo_bytes: bytes, nome_arquivo: str, log: list) -> pd.DataFrame | None:
    try:
        ext = os.path.splitext(nome_arquivo)[1].lower()
        if ext in ('.xlsx', '.xls'):
            df = pd.read_excel(io.BytesIO(arquivo_bytes), dtype=str)
        else:
            raw_str = arquivo_bytes.decode('latin-1', errors='replace')
            sep     = ';' if raw_str.count(';') >= raw_str.count(',') else ','
            df      = pd.read_csv(io.StringIO(raw_str), sep=sep, dtype=str)
        df.columns = [str(c).strip().upper() for c in df.columns]
        colunas_obrigatorias = [
            'NF', 'COD.ITEM',
            'PIS CST', 'PIS PC ALIQ', 'PIS VL BASE', 'PIS VALOR',
            'COFINS CST', 'COFINS PC ALIQ', 'COFINS VL BASE', 'COFINS VALOR',
        ]
        ausentes = [c for c in colunas_obrigatorias if c not in df.columns]
        if ausentes:
            log.append(f"ERRO Planilha Cliente: colunas ausentes → {ausentes}")
            log.append(f"Colunas encontradas: {list(df.columns)}")
            return None
        df['NF']       = df['NF'].apply(lambda x: str(x).strip().split('.')[0])
        df['COD.ITEM'] = df['COD.ITEM'].apply(lambda x: str(x).strip().split('.')[0])
        # Normaliza DT EMISSAO se presente
        if 'DT EMISSAO' in df.columns:
            df['DT EMISSAO'] = df['DT EMISSAO'].apply(converter_data_planilha)
        tem_cbenef = 'CBENEF' in df.columns
        log.append(
            f"Planilha Cliente carregada: {len(df)} linhas | "
            f"NFs únicas: {df['NF'].nunique()} | "
            f"Itens únicos: {df['COD.ITEM'].nunique()} | "
            f"CBENEF: {'✔ presente' if tem_cbenef else '✘ ausente'}"
        )
        return df
    except Exception as e:
        log.append(f"ERRO ao carregar Planilha Cliente: {e}")
        log.append(traceback.format_exc())
        return None


def _safe_str(valor, casas: int = 2) -> str:
    s = str(valor).strip()
    if not s or s.upper() == 'NAN': return f"0,{'0' * casas}"
    s = s.replace(',', '.')
    try: return f"{float(s):.{casas}f}".replace('.', ',')
    except ValueError: return f"0,{'0' * casas}"


def _safe_int(valor) -> str:
    s = str(valor).strip()
    if not s or s.upper() == 'NAN': return ''
    return s.split('.')[0]


def _safe_cbenef(valor) -> str:
    s = str(valor).strip().replace(' ', '').upper()
    if not s or s == 'NAN': return ''
    return s[:60]  # campo 91 do 0100 = Identificador, máx 60 chars


def _safe_num(valor, default: str = '0') -> str:
    """Garante que o valor seja numérico inteiro (sem letras)."""
    s = str(valor).strip().split('.')[0]
    if s.lstrip('-').isdigit(): return s
    return default


def _safe_date_dominio(valor) -> str:
    """Converte qualquer data para DD/MM/AAAA (formato Domínio)."""
    s = str(valor).strip()
    if not s or s.upper() == 'NAN': return ''
    # Já está no formato correto
    if len(s) == 10 and s[2] == '/' and s[5] == '/': return s
    # Formato DDMMAAAA (sem separadores)
    d = s.replace('/', '').replace('-', '')
    if len(d) == 8 and d.isdigit():
        return f"{d[0:2]}/{d[2:4]}/{d[4:8]}"
    # Tenta parse pandas
    return converter_data_planilha(valor)


def buscar_pis_cofins_planilha(
    df_cliente: pd.DataFrame, num_nf: str, cod_item: str, log: list,
) -> dict | None:
    nf_norm   = str(num_nf).strip().split('.')[0]
    item_norm = str(cod_item).strip().split('.')[0]
    mascara   = (df_cliente['NF'] == nf_norm) & (df_cliente['COD.ITEM'] == item_norm)
    linhas    = df_cliente[mascara]
    if linhas.empty:
        log.append(f"  AVISO: NF={nf_norm} / Item={item_norm} não encontrado na Planilha Cliente.")
        return None
    row = linhas.iloc[0]
    def _col(nome: str, default: str = '') -> str:
        return str(row[nome]).strip() if nome in df_cliente.columns else default
    return {
        'pis_cst':      _safe_int(_col('PIS CST')),
        'pis_base':     _safe_str(_col('PIS VL BASE'), 2),
        'pis_aliq':     _safe_str(_col('PIS PC ALIQ'), 4),
        'pis_valor':    _safe_str(_col('PIS VALOR'),   2),
        'cofins_cst':   _safe_int(_col('COFINS CST')),
        'cofins_base':  _safe_str(_col('COFINS VL BASE'), 2),
        'cofins_aliq':  _safe_str(_col('COFINS PC ALIQ'), 4),
        'cofins_valor': _safe_str(_col('COFINS VALOR'),   2),
        'cbs_class':    _safe_int(_col('CLASS TRIB CBS')),
        'cbs_base':     _safe_str(_col('CBS BASE'),  2),
        'cbs_aliq':     _safe_str(_col('CBS ALIQ'),  2),
        'cbs_valor':    _safe_str(_col('CBS VALOR'), 2),
        'ibs_class':    _safe_int(_col('CLASS TRIB IBS')),
        'ibs_base':     _safe_str(_col('IBS BASE'),  2),
        'ibs_aliq':     _safe_str(_col('IBS ALIQ'),  2),
        'ibs_valor':    _safe_str(_col('IBS VALOR'), 2),
        'cbenef':       _safe_cbenef(_col('CBENEF')),
        # Campos extras da planilha para enriquecer 0100
        'ncm':          _col('NCM'),
        'cest':         _safe_int(_col('CEST ORI')),
        'cfop':         _col('CFOP'),
        'dt_emissao':   _safe_date_dominio(_col('DT EMISSAO')),
        'vl_produto':   _safe_str(_col('VLR PRODUTO'), 2),
        'vl_total_nf':  _safe_str(_col('VLR TOTAL NF'), 2),
        'ipi_valor':    _safe_str(_col('IPI VALOR'), 2),
        'icms_valor':   _safe_str(_col('ICMS NORMAL VLR'), 2),
    }


def extrair_cbenef_por_produto(df_cliente: pd.DataFrame | None, log: list) -> dict:
    resultado = {}
    if df_cliente is None or df_cliente.empty: return resultado
    if 'CBENEF' not in df_cliente.columns:
        log.append("INFO: Coluna CBENEF não encontrada na Planilha Cliente.")
        return resultado
    for _, row in df_cliente.iterrows():
        cod_item = str(row.get('COD.ITEM', '')).strip().split('.')[0]
        cbenef   = _safe_cbenef(str(row.get('CBENEF', '')))
        if cod_item and cod_item not in resultado and cbenef:
            resultado[cod_item] = cbenef
    log.append(f"CBENEF extraído: {len([v for v in resultado.values() if v])} produto(s).")
    return resultado


def extrair_info_produto_da_planilha(df_cliente: pd.DataFrame | None, log: list) -> dict:
    """
    Extrai informações de produto da Planilha Cliente para enriquecer o cadastro 0100/0110.
    Chave: COD.ITEM → dict com NCM, CEST, CST PIS entrada, CST PIS saída, alíquotas, etc.
    """
    resultado = {}
    if df_cliente is None or df_cliente.empty:
        return resultado

    def _col(row, nome: str, default: str = '') -> str:
        return str(row[nome]).strip() if nome in df_cliente.columns else default

    for _, row in df_cliente.iterrows():
        cod_item = str(row.get('COD.ITEM', '')).strip().split('.')[0]
        if not cod_item or cod_item in resultado:
            continue

        pis_cst_entrada  = _safe_int(_col(row, 'PIS CST'))
        cofins_cst_ent   = _safe_int(_col(row, 'COFINS CST'))
        pis_aliq         = _safe_str(_col(row, 'PIS PC ALIQ'), 4)
        cofins_aliq      = _safe_str(_col(row, 'COFINS PC ALIQ'), 4)
        cst_pis_saida    = _cst_entrada_para_saida(pis_cst_entrada)
        cst_cofins_saida = _cst_entrada_para_saida(cofins_cst_ent)

        # Natureza de receita: busca pelo CST de saída
        nat_receita = ''
        if cst_pis_saida in CSTS_NATUREZA_OBRIGATORIA:
            nat_receita = ''  # será preenchida pela config do usuário

        resultado[cod_item] = {
            'ncm':              _col(row, 'NCM'),
            'cest':             _safe_int(_col(row, 'CEST ORI')),
            'cst_pis_entrada':  pis_cst_entrada,
            'cst_cof_entrada':  cofins_cst_ent,
            'cst_pis_saida':    cst_pis_saida,
            'cst_cof_saida':    cst_cofins_saida,
            'aliq_pis':         pis_aliq,
            'aliq_cofins':      cofins_aliq,
            'cbenef':           _safe_cbenef(_col(row, 'CBENEF')),
        }

    log.append(f"Info de produto extraída da Planilha Cliente: {len(resultado)} produto(s).")
    return resultado


def extrair_csts_obrigatorios_da_planilha(df_cliente: pd.DataFrame | None, log: list) -> set:
    """
    Coleta CSTs de PIS e COFINS da planilha, converte CSTs de entrada para saída,
    e retorna APENAS os CSTs de SAÍDA que possuem natureza OBRIGATÓRIA.
    """
    if df_cliente is None or df_cliente.empty:
        return set()

    todos_csts_saida = set()
    for col in ('PIS CST', 'COFINS CST'):
        if col not in df_cliente.columns:
            continue
        for val in df_cliente[col].dropna():
            s = str(val).strip().split('.')[0]
            try:
                cst_norm = str(int(s)).zfill(2)
            except ValueError:
                if s: cst_norm = s.zfill(2)
                else: continue
            # Converte para CST de saída
            try:
                n = int(cst_norm)
            except ValueError:
                n = 0
            if n >= 50:
                cst_saida = CST_ENTRADA_PARA_SAIDA.get(cst_norm, cst_norm)
            else:
                cst_saida = cst_norm
            todos_csts_saida.add(cst_saida)

    csts_obrig = todos_csts_saida & CSTS_NATUREZA_OBRIGATORIA
    csts_opcio = todos_csts_saida - CSTS_NATUREZA_OBRIGATORIA

    log.append(
        f"CSTs de saída identificados: {sorted(todos_csts_saida)} | "
        f"Obrigatórios: {sorted(csts_obrig)} | "
        f"Opcionais (não exibidos): {sorted(csts_opcio)}"
    )
    return csts_obrig


# ==============================
# PARSER SPED FISCAL
# ==============================
def parse_sped(content: str) -> dict:
    linhas_ordenadas = []; por_tipo = {}
    for num_linha, linha in enumerate(content.splitlines(), start=1):
        linha = linha.strip()
        if not linha: continue
        campos = linha.split('|')
        if campos and campos[0] == '':  campos = campos[1:]
        if campos and campos[-1] == '': campos = campos[:-1]
        if not campos: continue
        tipo = campos[0].strip()
        if not tipo: continue
        linhas_ordenadas.append((tipo, campos, num_linha))
        por_tipo.setdefault(tipo, []).append((campos, num_linha))
    return {'linhas_ordenadas': linhas_ordenadas, 'por_tipo': por_tipo}


def _c(campos: list, idx: int, default: str = '') -> str:
    return campos[idx].strip() if len(campos) > idx else default


SPED_0000_CNPJ=6; SPED_0000_DT_INI=3
SPED_0150_COD=1; SPED_0150_CNPJ=4
SPED_0200_COD_ITEM=1; SPED_0200_DESCR=2; SPED_0200_UNID_INV=5
SPED_0200_TIPO_ITEM=6; SPED_0200_COD_NCM=7; SPED_0200_ALIQ_ICMS=11
SPED_C100_IND_OPER=1; SPED_C100_COD_PART=3; SPED_C100_COD_MOD=4
SPED_C100_COD_SIT=5; SPED_C100_SER=6; SPED_C100_NUM_DOC=7
SPED_C100_CHV_NFE=8; SPED_C100_DT_DOC=9; SPED_C100_DT_ES=10
SPED_C100_VL_DOC=11; SPED_C100_VL_ICMS=21
SPED_C170_NUM_ITEM=1; SPED_C170_COD_ITEM=2; SPED_C170_DESCR=3
SPED_C170_QTD=4; SPED_C170_UNID=5; SPED_C170_VL_ITEM=6
SPED_C170_VL_DESC=7; SPED_C170_CST_ICMS=9; SPED_C170_CFOP=10
SPED_C170_VL_BC=12; SPED_C170_ALIQ_ICMS=13; SPED_C170_VL_ICMS=14
SPED_C170_VL_BC_ST=15; SPED_C170_ALIQ_ST=16; SPED_C170_VL_ICMS_ST=17
SPED_C170_CST_IPI=19; SPED_C170_VL_BC_IPI=21
SPED_C170_ALIQ_IPI=22; SPED_C170_VL_IPI=23
SPED_C170_CST_PIS=24; SPED_C170_VL_BC_PIS=25; SPED_C170_ALIQ_PIS=26
SPED_C170_VL_PIS=28; SPED_C170_CST_COFINS=29; SPED_C170_VL_BC_COF=30
SPED_C170_ALIQ_COF=31; SPED_C170_VL_COFINS=33
SPED_C190_CFOP=2; SPED_C190_ALIQ=3
SPED_D100_IND_OPER=1; SPED_D100_COD_PART=3; SPED_D100_COD_MOD=4
SPED_D100_COD_SIT=5; SPED_D100_SER=6; SPED_D100_NUM_DOC=8
SPED_D100_DT_DOC=10; SPED_D100_VL_DOC=14; SPED_D100_ALIQ=19; SPED_D100_VL_ICMS=20
SPED_H010_COD_ITEM=1; SPED_H010_UNID=2; SPED_H010_QTD=3
SPED_H010_VL_UNIT=4; SPED_H010_VL_ITEM=5


def extrair_cfops_do_sped(parsed: dict, tabela_flags: dict, log: list) -> dict:
    cfops = {}
    def registrar(cfop_raw: str, tipo_reg: str):
        cfop = str(cfop_raw).strip()
        if not cfop or not cfop.isdigit(): return
        cfop = cfop.zfill(4)
        if cfop == '0000': return
        if cfop not in cfops:
            flags = get_flags_cfop(cfop, tabela_flags)
            cfops[cfop] = {
                'registros': set(), 'ocorrencias': 0,
                'tipo_operacao': get_tipo_operacao(cfop),
                'indNFe': flags['indNFe'], 'indComunica': flags['indComunica'],
                'indTransp': flags['indTransp'], 'indDevol': flags['indDevol'],
            }
        cfops[cfop]['registros'].add(tipo_reg)
        cfops[cfop]['ocorrencias'] += 1
    contadores = {'C100': 0, 'C170': 0, 'C190': 0, 'D100': 0}
    for tipo, campos, _ in parsed['linhas_ordenadas']:
        if tipo in contadores: contadores[tipo] += 1
        if tipo == 'C170':   registrar(_c(campos, SPED_C170_CFOP), 'C170')
        elif tipo == 'C190': registrar(_c(campos, SPED_C190_CFOP), 'C190')
    log.append(f"Registros lidos: C100={contadores['C100']} | C170={contadores['C170']} | "
               f"C190={contadores['C190']} | D100={contadores['D100']}")
    log.append(f"CFOPs únicos encontrados: {len(cfops)} — {sorted(cfops.keys())}")
    return cfops


def extrair_produtos_do_sped(parsed: dict, log: list) -> dict:
    produtos = {}
    if '0200' in parsed['por_tipo']:
        for campos, _ in parsed['por_tipo']['0200']:
            cod = _c(campos, SPED_0200_COD_ITEM).strip()
            if not cod: continue
            produtos[cod] = {
                'descr': _c(campos, SPED_0200_DESCR), 'unid': _c(campos, SPED_0200_UNID_INV),
                'ncm': _c(campos, SPED_0200_COD_NCM), 'tipo_item': _c(campos, SPED_0200_TIPO_ITEM),
                'aliq_icms': _c(campos, SPED_0200_ALIQ_ICMS) or '0,00', 'aliq_ipi': '0,00',
                'cst_icms': '', 'cst_ipi': '', 'cst_pis': '', 'cst_cofins': '',
                'aliq_pis': '0,00', 'aliq_cofins': '0,00', 'cest': '',
            }
        log.append(f"Produtos carregados do 0200: {len(produtos)}")
    itens_c170 = 0
    for tipo, campos, _ in parsed['linhas_ordenadas']:
        if tipo != 'C170': continue
        cod = _c(campos, SPED_C170_COD_ITEM).strip()
        if not cod: continue
        itens_c170 += 1
        a_icms = _c(campos, SPED_C170_ALIQ_ICMS) or '0,00'
        a_ipi  = _c(campos, SPED_C170_ALIQ_IPI)  or '0,00'
        a_pis  = _c(campos, SPED_C170_ALIQ_PIS)  or '0,00'
        a_cof  = _c(campos, SPED_C170_ALIQ_COF)  or '0,00'
        c_icms = _c(campos, SPED_C170_CST_ICMS)
        c_ipi  = _c(campos, SPED_C170_CST_IPI)
        c_pis  = _c(campos, SPED_C170_CST_PIS)
        c_cof  = _c(campos, SPED_C170_CST_COFINS)
        if cod not in produtos:
            produtos[cod] = {
                'descr': _c(campos, SPED_C170_DESCR), 'unid': _c(campos, SPED_C170_UNID),
                'ncm': '', 'tipo_item': '00', 'aliq_icms': a_icms, 'aliq_ipi': a_ipi,
                'cst_icms': c_icms, 'cst_ipi': c_ipi, 'cst_pis': c_pis, 'cst_cofins': c_cof,
                'aliq_pis': a_pis, 'aliq_cofins': a_cof, 'cest': '',
            }
        else:
            p = produtos[cod]
            if not p.get('cst_icms'):   p['cst_icms']   = c_icms
            if not p.get('cst_ipi'):    p['cst_ipi']    = c_ipi
            if not p.get('cst_pis'):    p['cst_pis']    = c_pis
            if not p.get('cst_cofins'): p['cst_cofins'] = c_cof
            for k, v in [('aliq_ipi', a_ipi), ('aliq_pis', a_pis),
                         ('aliq_cofins', a_cof), ('aliq_icms', a_icms)]:
                if p.get(k, '0,00') in ('', '0,00', '0'): p[k] = v
            if not p.get('unid'):  p['unid']  = _c(campos, SPED_C170_UNID)
            if not p.get('descr'): p['descr'] = _c(campos, SPED_C170_DESCR)
    log.append(f"Itens C170 processados: {itens_c170} | Total produtos únicos: {len(produtos)}")
    return produtos


def gerar_xlsx_acumuladores_tr(cfops_dict: dict, tabela_descr: dict, tabela_flags: dict) -> bytes:
    wb = Workbook()
    COR_LARANJA="FF8000"; COR_CINZA_ESC="444444"; COR_CINZA_CLR="E9E9E9"
    COR_BRANCO="FFFFFF"; COR_LARANJA_C="FFF3E0"; COR_VERM_CLR="FFEBEE"
    borda_fina = Border(
        left=Side(style='thin', color="CCCCCC"), right=Side(style='thin', color="CCCCCC"),
        top=Side(style='thin',  color="CCCCCC"), bottom=Side(style='thin', color="CCCCCC"),
    )
    def fill(h): return PatternFill("solid", fgColor=h)
    def center(): return Alignment(horizontal='center', vertical='center')
    def left_al(): return Alignment(horizontal='left', vertical='center')
    def wrap_al(): return Alignment(horizontal='left', vertical='center', wrap_text=True)
    cfops_ord = sorted(cfops_dict.items(),
                       key=lambda x: (0 if x[1]['tipo_operacao'] == 'Entrada' else 1, x[0]))
    ws1 = wb.active; ws1.title = "Acumuladores"; ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:I1"); ws1.row_dimensions[1].height = 36
    c = ws1["A1"]; c.value = "Thomson Reuters  |  Domínio Sistemas  —  Tabela de Acumuladores CFOP"
    c.fill = fill(COR_CINZA_ESC); c.font = Font(name='Segoe UI', bold=True, size=13, color=COR_LARANJA); c.alignment = left_al()
    ws1.merge_cells("A2:I2"); ws1.row_dimensions[2].height = 20
    c2 = ws1["A2"]
    c2.value = (f"Extraído do SPED Fiscal  |  {len(cfops_dict)} CFOP(s)  |  "
                f"Descrições: Receita Federal  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c2.fill = fill(COR_LARANJA); c2.font = Font(name='Segoe UI', size=9, color=COR_BRANCO); c2.alignment = left_al()
    ws1.merge_cells("A3:I3"); ws1.row_dimensions[3].height = 18
    c3 = ws1["A3"]; c3.value = "⚠  Preencha a coluna ACUMULADOR para cada CFOP antes de fazer o upload no conversor."
    c3.fill = fill(COR_CINZA_CLR); c3.font = Font(name='Segoe UI', bold=True, size=9, color=COR_CINZA_ESC); c3.alignment = left_al()
    ws1.row_dimensions[4].height = 6; ws1.row_dimensions[5].height = 22
    cabecalhos = ['CFOP', 'DESCRIÇÃO OFICIAL (Receita Federal)', 'TIPO OPERAÇÃO',
                  'OCORRÊNCIAS', 'DEVOLUÇÃO', 'NFe', 'COMUNICAÇÃO', 'TRANSPORTE', 'ACUMULADOR']
    col_widths = [10, 60, 16, 14, 12, 8, 14, 14, 16]
    for ci, (cab, w) in enumerate(zip(cabecalhos, col_widths), start=1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
        cell = ws1.cell(row=5, column=ci, value=cab)
        cell.fill = fill(COR_LARANJA); cell.font = Font(name='Segoe UI', bold=True, size=11, color=COR_BRANCO)
        cell.alignment = center(); cell.border = borda_fina
    linha = 6
    for idx, (cfop, info) in enumerate(cfops_ord):
        ws1.row_dimensions[linha].height = 30
        bg = (COR_CINZA_CLR if idx % 2 == 0 else COR_BRANCO) if info['tipo_operacao'] == 'Entrada' \
             else (COR_LARANJA_C if idx % 2 == 0 else COR_BRANCO)
        descricao = get_descricao_cfop(cfop, tabela_descr)
        eh_devol = '✔' if info.get('indDevol') == 1 else ''
        eh_nfe   = '✔' if info.get('indNFe')   == 1 else ''
        eh_com   = '✔' if info.get('indComunica') == 1 else ''
        eh_trp   = '✔' if info.get('indTransp')   == 1 else ''
        valores  = [cfop, descricao, info['tipo_operacao'], info['ocorrencias'],
                    eh_devol, eh_nfe, eh_com, eh_trp, '']
        for ci, valor in enumerate(valores, start=1):
            cell = ws1.cell(row=linha, column=ci, value=valor); cell.border = borda_fina
            if ci == 1:
                cell.fill = fill(bg); cell.font = Font(name='Segoe UI', bold=True, size=10, color=COR_CINZA_ESC); cell.alignment = center()
            elif ci == 2:
                cell.fill = fill(bg); cell.font = Font(name='Segoe UI', size=9, color=COR_CINZA_ESC); cell.alignment = wrap_al()
            elif ci == 3:
                cell.fill = fill(bg); cor = "1B5E20" if info['tipo_operacao'] == 'Entrada' else "B71C1C"
                cell.font = Font(name='Segoe UI', bold=True, size=10, color=cor); cell.alignment = center()
            elif ci == 4:
                cell.fill = fill(bg); cell.font = Font(name='Segoe UI', size=10, color=COR_CINZA_ESC); cell.alignment = center()
            elif ci in (5, 6, 7, 8):
                cor_flag = "B71C1C" if (ci == 5 and valor == '✔') else ("1B5E20" if valor == '✔' else COR_CINZA_ESC)
                cell.fill = fill(COR_VERM_CLR if ci == 5 and valor == '✔' else bg)
                cell.font = Font(name='Segoe UI', bold=True, size=10, color=cor_flag); cell.alignment = center()
            elif ci == 9:
                cell.fill = fill("FFF8F0"); cell.font = Font(name='Segoe UI', bold=True, size=10, color=COR_LARANJA)
                cell.alignment = center()
                cell.border = Border(left=Side(style='medium', color=COR_LARANJA),
                                     right=Side(style='medium', color=COR_LARANJA),
                                     top=Side(style='thin', color="CCCCCC"),
                                     bottom=Side(style='thin', color="CCCCCC"))
        linha += 1
    ws1.merge_cells(f"A{linha}:I{linha}"); ws1.row_dimensions[linha].height = 18
    cr = ws1.cell(row=linha, column=1, value="Thomson Reuters  |  Domínio Sistemas  |  Descrições: Receita Federal")
    cr.fill = fill(COR_CINZA_ESC); cr.font = Font(name='Segoe UI', size=8, color="888888")
    cr.alignment = Alignment(horizontal='right', vertical='center')
    ws1.freeze_panes = "A6"; ws1.auto_filter.ref = f"A5:I{linha - 1}"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def carregar_acumuladores(arquivo_bytes: bytes, nome_arquivo: str, log: list) -> dict:
    try:
        ext = os.path.splitext(nome_arquivo)[1].lower()
        if ext in ('.xlsx', '.xls'):
            df = None
            for header_row in range(6):
                try:
                    df_tent = pd.read_excel(io.BytesIO(arquivo_bytes), sheet_name=0,
                                            header=header_row, dtype=str)
                    cols = [str(c).strip().upper() for c in df_tent.columns]
                    if 'CFOP' in cols and 'ACUMULADOR' in cols:
                        df_tent.columns = cols; df = df_tent
                        log.append(f"Planilha lida (header linha {header_row + 1}). Colunas: {list(df.columns)}")
                        break
                except Exception: continue
            if df is None:
                log.append("ERRO: Não foi possível localizar as colunas 'CFOP' e 'ACUMULADOR'."); return None
        else:
            raw_str = arquivo_bytes.decode('latin-1', errors='replace')
            sep = ';' if raw_str.count(';') >= raw_str.count(',') else ','
            df  = pd.read_csv(io.StringIO(raw_str), sep=sep, dtype=str)
            df.columns = [str(c).strip().upper() for c in df.columns]
            if 'CFOP' not in df.columns or 'ACUMULADOR' not in df.columns:
                log.append("ERRO: CSV deve conter as colunas 'CFOP' e 'ACUMULADOR'."); return None
        tabela = {}
        for _, row in df.iterrows():
            raw_cfop = str(row['CFOP']).strip()
            if '.' in raw_cfop: raw_cfop = raw_cfop.split('.')[0]
            raw_cfop = ''.join(filter(str.isdigit, raw_cfop))
            if not raw_cfop: continue
            cfop = raw_cfop.zfill(4)
            raw_acum = str(row['ACUMULADOR']).strip()
            if raw_acum.endswith('.0'): raw_acum = raw_acum[:-2]
            acum = raw_acum.strip()
            if cfop == '0000': continue
            if not acum or acum.upper() in ('NAN', '', '0'): continue
            tabela[cfop] = acum
        if not tabela:
            log.append("ERRO: Nenhum par CFOP → Acumulador preenchido."); return None
        log.append(f"Tabela de acumuladores carregada: {len(tabela)} CFOPs mapeados.")
        return tabela
    except Exception as e:
        log.append(f"ERRO ao carregar tabela de acumuladores: {e}")
        log.append(traceback.format_exc()); return None


def get_acumulador(cfop: str, tabela: dict, nao_mapeados: set) -> str:
    cfop_norm = str(cfop).strip().zfill(4)
    acum = tabela.get(cfop_norm)
    if acum is None: nao_mapeados.add(cfop_norm); return '9999'
    return acum


# ==============================
# GERAÇÃO DOS REGISTROS 0100 + 0110
# V3.3 — enriquecido com dados da Planilha Cliente
# Leiaute 0100: campo 91 = Identificador (CBENEF, máx 60 chars)
# Leiaute 0110: campo 16 = CST Saídas | campo 18 = Natureza de receita
# ==============================
def gerar_registros_produtos(
    produtos: dict,
    dt_ini: str,
    por_nome_imp: dict,
    log: list,
    cbenef_por_produto: dict | None = None,
    info_produto_planilha: dict | None = None,
    config_natureza: dict | None = None,
) -> str:
    saida = StringIO(); n_prod = 0; n_vig = 0
    cb_map   = cbenef_por_produto or {}
    info_map = info_produto_planilha or {}
    cfg_nat  = config_natureza or {}

    for cod, p in sorted(produtos.items()):
        descr      = (p.get('descr') or '').replace('|', ' ')[:60]
        unid       = p.get('unid', 'UN') or 'UN'
        aliq_icms  = p.get('aliq_icms', '0,00') or '0,00'
        aliq_ipi   = p.get('aliq_ipi',  '0,00') or '0,00'
        cst_icms   = p.get('cst_icms',  '') or ''
        cst_ipi    = p.get('cst_ipi',   '') or ''
        cst_pis    = p.get('cst_pis',   '') or ''
        cst_cofins = p.get('cst_cofins','') or ''
        aliq_pis   = p.get('aliq_pis',  '0,00') or '0,00'
        aliq_cof   = p.get('aliq_cofins','0,00') or '0,00'

        # Enriquece com dados da Planilha Cliente
        info_pl = info_map.get(cod, {})
        ncm     = info_pl.get('ncm', '') or p.get('ncm', '') or ''
        cest    = info_pl.get('cest', '') or p.get('cest', '') or ''

        # CSTs da planilha têm prioridade
        if info_pl.get('cst_pis_entrada'):
            cst_pis    = info_pl['cst_pis_entrada']
        if info_pl.get('cst_cof_entrada'):
            cst_cofins = info_pl['cst_cof_entrada']
        if info_pl.get('aliq_pis') and info_pl['aliq_pis'] not in ('0,0000', '0,00'):
            aliq_pis = info_pl['aliq_pis']
        if info_pl.get('aliq_cofins') and info_pl['aliq_cofins'] not in ('0,0000', '0,00'):
            aliq_cof = info_pl['aliq_cofins']

        # CST de saída (De-Para)
        cst_pis_saida    = info_pl.get('cst_pis_saida',  _cst_entrada_para_saida(cst_pis))
        cst_cofins_saida = info_pl.get('cst_cof_saida',  _cst_entrada_para_saida(cst_cofins))

        # Natureza de receita (campo 18 do 0110) — busca pelo CST de saída
        nat_receita = ''
        chave_nat = f"cst_{cst_pis_saida}_nat_pis"
        if chave_nat in cfg_nat and cfg_nat[chave_nat]:
            nat_receita = cfg_nat[chave_nat]
        elif f"cst_{cst_cofins_saida}_nat_cofins" in cfg_nat:
            nat_receita = cfg_nat[f"cst_{cst_cofins_saida}_nat_cofins"]

        # CBENEF → campo 91 do 0100 (Identificador, máx 60 chars)
        cbenef = info_pl.get('cbenef', '') or cb_map.get(cod, '')

        # ── Registro 0100 ─────────────────────────────────────────────────
        # Campos 1-27 conforme leiaute fornecido
        # Campos 28-90 = campos estaduais/específicos (vazios)
        # Campo 91 = Identificador (CBENEF)
        saida.write(
            f"|0100|{cod}|{descr}|||{ncm}||||{unid}|N|O|||"
            f"|N||0,000|0,00000|0,000|{cst_icms}|{aliq_icms}|{aliq_ipi}|M||N|"
            f"{'|' * 63}"   # campos 28-90 = 63 campos vazios
            f"{cest}|"      # campo 89 = CEST
            f"|"            # campo 90 = Registro de Exportação (RE)
            f"{cbenef}|\n"  # campo 91 = Identificador (CBENEF)
        )
        n_prod += 1

        # ── Registro 0110 ─────────────────────────────────────────────────
        # Leiaute 0110 fornecido:
        # Campo 1  = 0110 | Campo 2 = Descrição | Campo 3 = CST Entrada
        # Campo 4  = Vínculo do Crédito | Campo 5 = Base do Crédito
        # Campo 6  = Aproveitar crédito proporcional (N)
        # Campo 7  = Crédito por alíquota diferenciada Entradas (N)
        # Campo 8  = Alíquota PIS Entradas | Campo 9 = Alíquota COFINS Entradas
        # Campo 10 = Crédito por unidade de medida (N)
        # Campo 11 = Unidade tributada diferente (N)
        # Campo 12 = Unidade tributável | Campo 13 = Fator de conversão
        # Campo 14 = Valor PIS Entradas | Campo 15 = Valor COFINS Entradas
        # Campo 16 = CST Saídas ← De-Para do CST de entrada
        # Campo 17 = Tipo de contribuição (N=Não cumulativo, C=Cumulativo, S=Sem incidência)
        # Campo 18 = Natureza de receita ← preenchida pela config do usuário
        # Campo 19 = Código de recolhimento PIS Saída
        # Campo 20 = Código de recolhimento COFINS Saída
        # Campo 21 = Débito por alíquota diferenciada Saídas (N)
        # Campo 22 = Alíquota PIS Saídas | Campo 23 = Alíquota COFINS Saídas
        # Campo 24 = Débito por unidade de medida (N)
        # Campo 25 = Unidade tributada diferente Saídas (N)
        # Campo 26 = Unidade tributável Saídas | Campo 27 = Fator de conversão Saídas
        # Campo 28 = Valor PIS Saídas | Campo 29 = Valor COFINS Saídas
        # Campo 30 = Tabela SPED | Campo 31 = Marca/Grupo SPED
        # Campos 32-40 = outros campos conforme leiaute
        # Campos 34 = ICMS CST/CSOSN Entradas | Campo 35 = ICMS CST/CSOSN Saídas
        # Campo 36 = ICMS Alíquota | Campo 37 = IPI CST Entradas | Campo 38 = IPI CST Saídas
        # Campo 39 = IPI Periodicidade | Campo 40 = IPI Alíquota
        # Campos 41+ = Simples Nacional e outros

        # Determina tipo de contribuição pelo CST de saída
        tipo_contrib = 'N'  # Não cumulativo padrão
        try:
            n_cst_saida = int(cst_pis_saida)
            if n_cst_saida in (6, 7, 8, 9): tipo_contrib = 'S'  # Sem incidência
        except ValueError: pass

        saida.write(
            f"|0110|Vigência||01|N|N|"
            # campo 8=aliq_pis_ent | campo 9=aliq_cof_ent
            f"{aliq_pis}|{aliq_cof}|N|N|||"
            # campos 14-15 = valor PIS/COFINS entradas
            f"0,0000|0,0000|"
            # campo 16 = CST Saídas (De-Para)
            f"{cst_pis_saida}|"
            # campo 17 = Tipo de contribuição
            f"{tipo_contrib}|"
            # campo 18 = Natureza de receita
            f"{nat_receita}|"
            # campos 19-20 = cod recolhimento PIS/COFINS saída
            f"||"
            # campo 21 = débito alíq diferenciada (N)
            f"N|"
            # campos 22-23 = alíq PIS/COFINS saídas
            f"{aliq_pis}|{aliq_cof}|"
            # campo 24 = débito por unidade (N)
            f"N|N|||"
            # campos 28-29 = valor PIS/COFINS saídas
            f"0,0000|0,0000|||"
            # campos 32-33 = PIS/COFINS cumulativo (N)
            f"N|N|"
            # campo 34 = ICMS CST entrada | campo 35 = ICMS CST saída
            f"{cst_icms}|{cst_icms}|"
            # campo 36 = alíquota ICMS
            f"{aliq_icms}|"
            # campo 37 = IPI CST entrada | campo 38 = IPI CST saída
            f"{cst_ipi}|{cst_ipi}|"
            # campo 39 = periodicidade IPI | campo 40 = alíquota IPI
            f"M|{aliq_ipi}|"
            # campos 41+ = outros campos
            f"{'|' * 30}\n"
        )
        n_vig += 1

    log.append(
        f"Produtos gerados: {n_prod} registros 0100 + {n_vig} registros 0110 | "
        f"CBENEF: {len([v for v in cb_map.values() if v])} produto(s) | "
        f"Enriquecidos da planilha: {len(info_map)} produto(s)"
    )
    return saida.getvalue()


# ==============================
# WIDGET DE CONFIGURAÇÃO DE NATUREZA PIS/COFINS — V3.3
# Exibe APENAS CSTs de saída obrigatórios encontrados na Planilha Cliente.
# ==============================
def render_configuracao_natureza(csts_obrigatorios: set) -> dict:
    """
    Renderiza o widget de configuração de Natureza de Receita.
    Um único campo por CST preenche tanto PIS quanto COFINS
    (mesma regra para ambos, conforme solicitado).
    """
    st.markdown("### 🏷️ Natureza de PIS/COFINS — CSTs com Preenchimento Obrigatório")

    cfg_anterior = st.session_state.get('config_natureza', {})
    config = {}

    if csts_obrigatorios:
        csts_exibir = sorted(
            [c for c in csts_obrigatorios if c in NATUREZA_POR_CST],
            key=lambda x: int(x) if x.isdigit() else 999
        )
        planilha_carregada = True
    else:
        csts_exibir = sorted(
            NATUREZA_POR_CST.keys(),
            key=lambda x: int(x) if x.isdigit() else 999
        )
        planilha_carregada = False

    if planilha_carregada and csts_exibir:
        st.markdown(
            f"""<div style="background:#FFEBEE; border-left:4px solid #B71C1C;
            border-radius:4px; padding:10px 14px; margin:6px 0; font-size:13px;">
            ⚠️ <strong>CSTs com natureza obrigatória detectados:</strong>
            &nbsp;{' &nbsp;·&nbsp; '.join([f'<code>{c}</code>' for c in csts_exibir])}
            <br><small>Configure a Natureza de Receita abaixo.
            O mesmo código será aplicado a PIS e COFINS (campo 18 do 0110
            e campos 71/72 do 1030 e 2030).</small>
            </div>""",
            unsafe_allow_html=True,
        )
    elif planilha_carregada and not csts_exibir:
        st.markdown(
            """<div class="info-ok-box">
            ✅ <strong>Nenhum CST com natureza obrigatória detectado na Planilha Cliente.</strong><br>
            <small>Os CSTs encontrados não exigem Natureza de Receita.
            A conversão pode prosseguir normalmente.</small>
            </div>""",
            unsafe_allow_html=True,
        )
        st.session_state['config_natureza'] = config
        return config
    else:
        st.markdown(
            """<div class="natureza-box">
            ℹ️ Faça o upload da <strong>Planilha Cliente</strong> primeiro para que apenas
            os CSTs relevantes sejam exibidos aqui.
            </div>""",
            unsafe_allow_html=True,
        )

    if not csts_exibir:
        st.session_state['config_natureza'] = config
        return config

    # ── Tabela de referência ──────────────────────────────────────────────
    with st.expander("📊 Tabela de Naturezas de Receita (CSTs obrigatórios)", expanded=False):
        rows_tabela = []
        for cst_cod in csts_exibir:
            naturezas = NATUREZA_POR_CST.get(cst_cod, {})
            cst_descr = CST_DESCRICOES.get(cst_cod, f'CST {cst_cod}')
            for nat_cod, nat_descr in naturezas.items():
                rows_tabela.append({
                    'CST':                              cst_cod,
                    'Descrição CST':                    cst_descr,
                    'Cód. Natureza':                    nat_cod,
                    'Descrição da Natureza de Receita': nat_descr,
                })
        if rows_tabela:
            df_tabela = pd.DataFrame(rows_tabela)
            st.dataframe(df_tabela, use_container_width=True, hide_index=True, height=300)
            csv_tabela = df_tabela.to_csv(index=False, sep=';', encoding='utf-8-sig')
            st.download_button(
                label="⬇ Baixar Tabela (.csv)",
                data=csv_tabela.encode('utf-8-sig'),
                file_name="natureza_receita.csv",
                mime="text/csv",
                use_container_width=False,
            )

    st.markdown("---")

    opcoes_base_cods   = list(BASE_CREDITO_OPCOES.keys())
    opcoes_base_labels = list(BASE_CREDITO_OPCOES.values())
    opcoes_vinc_cods   = list(VINCULO_CREDITO_OPCOES.keys())
    opcoes_vinc_labels = list(VINCULO_CREDITO_OPCOES.values())

    with st.expander(f"⚙️ Configurar Natureza — {len(csts_exibir)} CST(s)", expanded=True):
        for cst_cod in csts_exibir:
            naturezas = NATUREZA_POR_CST.get(cst_cod, {})
            cst_descr = CST_DESCRICOES.get(cst_cod, f'CST {cst_cod}')

            st.markdown(
                f"<div class='cst-header-obrig'>{cst_descr} &nbsp;⚠ OBRIGATÓRIO</div>",
                unsafe_allow_html=True,
            )

            # ── Opções de natureza ────────────────────────────────────────
            nat_opcoes_cod   = [''] + list(naturezas.keys())
            nat_opcoes_label = ['-- Selecione a natureza --'] + [
                f"{k} – {v[:80]}{'...' if len(v) > 80 else ''}"
                for k, v in naturezas.items()
            ]

            # Chaves unificadas: nat usa uma só chave; base e vínculo também
            chave_nat  = f"cst_{cst_cod}_nat"       # ← único campo para PIS e COFINS
            chave_base = f"cst_{cst_cod}_base"
            chave_vinc = f"cst_{cst_cod}_vinc"

            # Retrocompatibilidade: lê chave antiga (nat_pis) se a nova não existir
            val_nat  = (cfg_anterior.get(chave_nat)
                        or cfg_anterior.get(f"cst_{cst_cod}_nat_pis", ''))
            val_base = (cfg_anterior.get(chave_base)
                        or cfg_anterior.get(f"cst_{cst_cod}_base_pis", ''))
            val_vinc = (cfg_anterior.get(chave_vinc)
                        or cfg_anterior.get(f"cst_{cst_cod}_vinc_pis", ''))

            idx_nat  = nat_opcoes_cod.index(val_nat)   if val_nat  in nat_opcoes_cod  else 0
            idx_base = opcoes_base_cods.index(val_base) if val_base in opcoes_base_cods else 0
            idx_vinc = opcoes_vinc_cods.index(val_vinc) if val_vinc in opcoes_vinc_cods else 0

            # ── Natureza (campo único — aplica a PIS e COFINS) ───────────
            st.markdown(
                "<span style='color:#FF8000;font-weight:bold;font-size:12px;'>"
                "🟠 Natureza de Receita — aplica a PIS e COFINS</span>",
                unsafe_allow_html=True,
            )
            sel_nat = st.selectbox(
                f"Natureza {cst_cod}",
                nat_opcoes_label,
                index=idx_nat,
                key=f"nat_{cst_cod}",
                label_visibility="collapsed",
                help="Código preenchido nos campos de Natureza de PIS e COFINS (mesmo valor para ambos).",
            )
            nat_val = nat_opcoes_cod[nat_opcoes_label.index(sel_nat)]

            # Propaga o mesmo valor para PIS e COFINS nas chaves de saída
            config[chave_nat]                        = nat_val   # chave nova (unificada)
            config[f"cst_{cst_cod}_nat_pis"]         = nat_val   # chave legada PIS
            config[f"cst_{cst_cod}_nat_cofins"]      = nat_val   # chave legada COFINS

            # ── Base de crédito e Vínculo (linha única) ──────────────────
            col_base, col_vinc = st.columns(2)
            with col_base:
                st.markdown(
                    "<span style='color:#888;font-size:11px;'>Base do Crédito</span>",
                    unsafe_allow_html=True,
                )
                sel_base = st.selectbox(
                    f"Base {cst_cod}",
                    opcoes_base_labels,
                    index=idx_base,
                    key=f"base_{cst_cod}",
                    label_visibility="collapsed",
                    help="Base do Crédito PIS/COFINS — campo 5 do 0110 / campo 67 do 1030",
                )
                base_val = opcoes_base_cods[opcoes_base_labels.index(sel_base)]
                config[chave_base]                       = base_val
                config[f"cst_{cst_cod}_base_pis"]        = base_val
                config[f"cst_{cst_cod}_base_cofins"]     = base_val

            with col_vinc:
                st.markdown(
                    "<span style='color:#888;font-size:11px;'>Vínculo do Crédito</span>",
                    unsafe_allow_html=True,
                )
                sel_vinc = st.selectbox(
                    f"Vínculo {cst_cod}",
                    opcoes_vinc_labels,
                    index=idx_vinc,
                    key=f"vinc_{cst_cod}",
                    label_visibility="collapsed",
                    help="Vínculo de Crédito PIS/COFINS — campo 4 do 0110 / campo 72 do 1030",
                )
                vinc_val = opcoes_vinc_cods[opcoes_vinc_labels.index(sel_vinc)]
                config[chave_vinc]                       = vinc_val
                config[f"cst_{cst_cod}_vinc_pis"]        = vinc_val
                config[f"cst_{cst_cod}_vinc_cofins"]     = vinc_val

            st.markdown("")

        # ── Resumo das configurações ativas ──────────────────────────────
        st.markdown("---")
        # Exibe apenas as chaves unificadas (evita duplicidade no resumo)
        configs_ativas = {
            k: v for k, v in config.items()
            if v and not k.endswith(('_pis', '_cofins'))  # oculta chaves legadas do resumo
        }
        if configs_ativas:
            st.markdown("**✅ Configurações ativas:**")
            cols_res = st.columns(3)
            tipo_map = {
                'nat':  'Natureza PIS+COFINS',
                'base': 'Base do Crédito',
                'vinc': 'Vínculo do Crédito',
            }
            for i, (chave, valor) in enumerate(configs_ativas.items()):
                partes = chave.split('_')   # ex: ['cst', '06', 'nat']
                cst_n  = partes[1]
                tipo   = partes[2]
                cols_res[i % 3].markdown(
                    f"<div class='natureza-ativa'>"
                    f"<strong>CST {cst_n} / {tipo_map.get(tipo, tipo)}</strong><br>"
                    f"<span style='color:#1B5E20;'>Código: {valor}</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
        else:
            st.warning("⚠ Nenhuma natureza configurada. Configure antes de converter.")

    st.session_state['config_natureza'] = config
    return config


# ==============================
# CONVERSÃO SPED FISCAL → DOMÍNIO SISTEMAS — V3.3
# Corrige erros do validador Domínio:
# - 1000: campos numéricos não podem ter letras
# - 1000: data da escrituração deve ser DD/MM/AAAA
# - 1030: datas no formato correto DD/MM/AAAA
# - 1030: campos numéricos limpos
# ==============================
def converter_sped_para_dominio(
    parsed: dict,
    tabela_acum: dict,
    tabela_flags: dict,
    por_nome_imp: dict,
    log: list,
    df_cliente: pd.DataFrame | None = None,
    config_natureza: dict | None = None,
    cbenef_por_produto: dict | None = None,
    info_produto_planilha: dict | None = None,
) -> tuple:
    saida        = StringIO()
    nao_mapeados = set()
    stats = {
        'nf_entrada': 0, 'nf_saida': 0, 'itens': 0,
        'analiticos': 0, 'transporte': 0, 'inventario': 0,
        'devolucoes': 0, 'produtos': 0, 'erros': 0,
        'itens_planilha': 0, 'itens_sped': 0,
        'natureza_aplicada': 0, 'cbenef_aplicado': 0,
    }
    cfg_nat  = config_natureza or {}
    cb_map   = cbenef_por_produto or {}
    info_map = info_produto_planilha or {}

    COD_ICMS    = get_codigo_imposto('ICMS',        por_nome_imp, 1)
    COD_IPI     = get_codigo_imposto('IPI',         por_nome_imp, 2)
    COD_ICMS_ST = get_codigo_imposto('ICMS RETIDO', por_nome_imp, 11)

    usa_planilha = df_cliente is not None and not df_cliente.empty
    log.append(f"Fonte PIS/COFINS/IBS/CBS: {'Planilha Cliente ✔' if usa_planilha else 'SPED Fiscal (fallback)'}")
    log.append(f"CBENEF: {len(cb_map)} produto(s) | Natureza: {len([v for v in cfg_nat.values() if v])} config(s)")

    dt_ini_sped = ''
    if '0000' in parsed['por_tipo']:
        campos_0000, _ = parsed['por_tipo']['0000'][0]
        dt_ini_sped = _c(campos_0000, SPED_0000_DT_INI)

    if '0000' in parsed['por_tipo']:
        campos, _ = parsed['por_tipo']['0000'][0]
        cnpj = _c(campos, SPED_0000_CNPJ)
        saida.write(f"|0000|{cnpj}|\n")
        log.append(f"0000: CNPJ={cnpj}")
    else:
        log.append("AVISO: Registro 0000 não encontrado.")

    participantes = {}
    if '0150' in parsed['por_tipo']:
        for campos, _ in parsed['por_tipo']['0150']:
            participantes[_c(campos, SPED_0150_COD)] = campos
        log.append(f"Participantes carregados: {len(participantes)}")

    produtos = extrair_produtos_do_sped(parsed, log)
    if produtos:
        saida.write(gerar_registros_produtos(
            produtos, dt_ini_sped, por_nome_imp, log,
            cbenef_por_produto=cb_map,
            info_produto_planilha=info_map,
            config_natureza=cfg_nat,
        ))
        stats['produtos'] = len(produtos)
    else:
        log.append("AVISO: Nenhum produto encontrado no SPED (0200/C170).")

    # Hierarquia C100 → C170 → C190
    blocos_c = []; bloco_atual = None
    for tipo, campos, _ in parsed['linhas_ordenadas']:
        if tipo == 'C100':
            if bloco_atual is not None: blocos_c.append(bloco_atual)
            bloco_atual = {'c100': campos, 'c170': [], 'c190': []}
        elif tipo == 'C170' and bloco_atual is not None: bloco_atual['c170'].append(campos)
        elif tipo == 'C190' and bloco_atual is not None: bloco_atual['c190'].append(campos)
    if bloco_atual is not None: blocos_c.append(bloco_atual)
    log.append(f"Blocos C100 montados: {len(blocos_c)}")

    for bloco in blocos_c:
        campos_c100 = bloco['c100']
        try:
            ind_oper = _c(campos_c100, SPED_C100_IND_OPER)
            cod_part = _c(campos_c100, SPED_C100_COD_PART)
            cod_mod  = _c(campos_c100, SPED_C100_COD_MOD)
            cod_sit  = _c(campos_c100, SPED_C100_COD_SIT)
            serie    = _c(campos_c100, SPED_C100_SER)
            num_doc  = _c(campos_c100, SPED_C100_NUM_DOC)
            chv_nfe  = _c(campos_c100, SPED_C100_CHV_NFE)
            dt_doc   = converter_data(_c(campos_c100, SPED_C100_DT_DOC))
            dt_es    = converter_data(_c(campos_c100, SPED_C100_DT_ES))
            vl_doc   = _c(campos_c100, SPED_C100_VL_DOC)
            vl_icms  = _c(campos_c100, SPED_C100_VL_ICMS)
            part_campos = participantes.get(cod_part, [])
            cnpj_part   = _c(part_campos, SPED_0150_CNPJ) if part_campos else ''
            tipo_es     = 'E' if ind_oper == '0' else 'S'
            cfop_principal = ''
            if bloco['c170']:   cfop_principal = _c(bloco['c170'][0], SPED_C170_CFOP)
            elif bloco['c190']: cfop_principal = _c(bloco['c190'][0], SPED_C190_CFOP)
            acum_principal = get_acumulador(cfop_principal, tabela_acum, nao_mapeados)
            eh_devol       = is_devolucao(cfop_principal, tabela_flags)
            if eh_devol: stats['devolucoes'] += 1
            aliq_icms_nf = '0,00'
            if bloco['c190']: aliq_icms_nf = _c(bloco['c190'][0], SPED_C190_ALIQ) or '0,00'
            obs = 'DEVOLUCAO' if eh_devol else ''

            # ── REGISTRO 1000 — CORRIGIDO ─────────────────────────────────
            # Erros do validador corrigidos:
            # - Campo 'CFOP estendido/detalhamento' (era 'C') → deve ser numérico → ''
            # - Campo 'Valor do Frete' (era 'S') → deve ser decimal → '0,00'
            # - Campo 'CFOP documento fiscal' (era 'S'/'E') → numérico → cfop_principal
            # - Campo 'Data da escrituração' (era '0,00') → data → dt_es
            # - Campo 'Ind. operação' (era 'S'/'E') → numérico → ind_oper
            saida.write(
                f"|1000|{num_doc}|{cnpj_part}||{ind_oper}|{cfop_principal}|"
                f"{serie}|{cod_mod}|{cod_sit}|{chv_nfe}|||"
                # campo 13=dt_doc | campo 14=dt_es (data escrituração — DD/MM/AAAA)
                f"{dt_doc}|{dt_es}|{vl_doc}||{obs}||||||||||{tipo_es}|"
                # campos 24-27 = valores zerados (decimal)
                f"0,00|0,00|0,00|0,00||0,00||||0,00|0,00|0,00||{vl_doc}|"
                # campos 38-39 = numérico (não pode ser 'C' ou 'S')
                f"0|0||||{acum_principal}||0,00||||||N|S||{tipo_es}||0|||||"
                # campos 50+ = vazios (numéricos ou caracteres, não letras em campos numéricos)
                f"||||||||||||0|{cod_sit}|0||0,00|0,00|0,00|||||||||||\n"
            )
            if ind_oper == '0': stats['nf_entrada'] += 1
            else:               stats['nf_saida']   += 1

            saida.write(
                f"|1020|{num_doc}||{vl_doc}|{aliq_icms_nf}|{vl_icms}|"
                f"0,00|0,00|0,00|0,00|{vl_doc}||||\n"
            )
            stats['analiticos'] += 1

            for campos_c170 in bloco['c170']:
                cod_item  = _c(campos_c170, SPED_C170_COD_ITEM)
                qtd       = _c(campos_c170, SPED_C170_QTD)     or '0'
                unid      = _c(campos_c170, SPED_C170_UNID)
                vl_item   = _c(campos_c170, SPED_C170_VL_ITEM) or '0,00'
                vl_desc_i = _c(campos_c170, SPED_C170_VL_DESC) or '0,00'
                cfop_item = _c(campos_c170, SPED_C170_CFOP)

                vl_bc_icms  = _c(campos_c170, SPED_C170_VL_BC)       or '0,00'
                aliq_icms   = _c(campos_c170, SPED_C170_ALIQ_ICMS)   or '0,00'
                vl_icms_i   = _c(campos_c170, SPED_C170_VL_ICMS)     or '0,00'
                vl_bc_st    = _c(campos_c170, SPED_C170_VL_BC_ST)    or '0,00'
                aliq_st     = _c(campos_c170, SPED_C170_ALIQ_ST)     or '0,00'
                vl_icms_st  = _c(campos_c170, SPED_C170_VL_ICMS_ST)  or '0,00'
                vl_bc_ipi   = _c(campos_c170, SPED_C170_VL_BC_IPI)   or '0,00'
                aliq_ipi    = _c(campos_c170, SPED_C170_ALIQ_IPI)    or '0,00'
                vl_ipi      = _c(campos_c170, SPED_C170_VL_IPI)      or '0,00'

                # PIS / COFINS / IBS / CBS — Planilha Cliente (prioridade)
                dados_pl = None
                if usa_planilha:
                    dados_pl = buscar_pis_cofins_planilha(df_cliente, num_doc, cod_item, log)

                if dados_pl is not None:
                    pis_cst      = dados_pl['pis_cst']
                    pis_base     = dados_pl['pis_base']
                    pis_aliq     = dados_pl['pis_aliq']
                    pis_valor    = dados_pl['pis_valor']
                    cofins_cst   = dados_pl['cofins_cst']
                    cofins_base  = dados_pl['cofins_base']
                    cofins_aliq  = dados_pl['cofins_aliq']
                    cofins_valor = dados_pl['cofins_valor']
                    ibs_class    = dados_pl['ibs_class']
                    ibs_base     = dados_pl['ibs_base']
                    ibs_aliq     = dados_pl['ibs_aliq']
                    ibs_valor    = dados_pl['ibs_valor']
                    cbs_class    = dados_pl['cbs_class']
                    cbs_base     = dados_pl['cbs_base']
                    cbs_aliq     = dados_pl['cbs_aliq']
                    cbs_valor    = dados_pl['cbs_valor']
                    cbenef_item  = dados_pl.get('cbenef', '') or cb_map.get(cod_item, '')
                    # Usa data da planilha se disponível
                    dt_doc_item  = dados_pl.get('dt_emissao', '') or dt_doc
                    if not dt_doc_item: dt_doc_item = dt_doc
                    cod_pis      = get_codigo_pis(pis_aliq, por_nome_imp)
                    cod_cofins   = get_codigo_cofins(cofins_aliq, por_nome_imp)
                    stats['itens_planilha'] += 1
                else:
                    pis_cst      = _c(campos_c170, SPED_C170_CST_PIS)
                    pis_base     = _c(campos_c170, SPED_C170_VL_BC_PIS)  or '0,00'
                    pis_aliq     = _c(campos_c170, SPED_C170_ALIQ_PIS)   or '0,00'
                    pis_valor    = _c(campos_c170, SPED_C170_VL_PIS)     or '0,00'
                    cofins_cst   = _c(campos_c170, SPED_C170_CST_COFINS)
                    cofins_base  = _c(campos_c170, SPED_C170_VL_BC_COF)  or '0,00'
                    cofins_aliq  = _c(campos_c170, SPED_C170_ALIQ_COF)   or '0,00'
                    cofins_valor = _c(campos_c170, SPED_C170_VL_COFINS)  or '0,00'
                    ibs_class = ibs_base = ibs_aliq = ibs_valor = ''
                    cbs_class = cbs_base = cbs_aliq = cbs_valor = ''
                    ibs_base = cbs_base = '0,00'; ibs_aliq = cbs_aliq = '0,00'
                    ibs_valor = cbs_valor = '0,00'
                    cbenef_item  = cb_map.get(cod_item, '')
                    dt_doc_item  = dt_doc
                    cod_pis      = get_codigo_pis(pis_aliq, por_nome_imp)
                    cod_cofins   = get_codigo_cofins(cofins_aliq, por_nome_imp)
                    stats['itens_sped'] += 1

                if cbenef_item: stats['cbenef_aplicado'] += 1

                # Natureza: usa CST de saída para buscar configuração
                nat_pis, nat_cofins, base_cred_pis, base_cred_cofins, vinc_pis, vinc_cofins = \
                    get_natureza_por_cst(pis_cst, cofins_cst, cfg_nat)
                if nat_pis or nat_cofins or base_cred_pis or vinc_pis:
                    stats['natureza_aplicada'] += 1

                # Garante que dt_doc_item está no formato correto DD/MM/AAAA
                dt_doc_item = _safe_date_dominio(dt_doc_item)
                if not dt_doc_item: dt_doc_item = dt_doc

                try:
                    vl_unit = f"{float(vl_item.replace(',','.')) / float(qtd.replace(',','.')):.3f}".replace('.', ',')
                except Exception:
                    vl_unit = vl_item

                def _val(s):
                    try: return float(str(s).replace(',', '.'))
                    except: return 0.0

                if _val(vl_icms_st) > 0:
                    cod_imp = COD_ICMS_ST; vl_bc_p = vl_bc_st; aliq_p = aliq_st; vl_imp_p = vl_icms_st
                elif _val(vl_icms_i) > 0:
                    cod_imp = COD_ICMS;    vl_bc_p = vl_bc_icms; aliq_p = aliq_icms; vl_imp_p = vl_icms_i
                elif _val(vl_ipi) > 0:
                    cod_imp = COD_IPI;     vl_bc_p = vl_bc_ipi;  aliq_p = aliq_ipi;  vl_imp_p = vl_ipi
                else:
                    cod_imp = COD_ICMS;    vl_bc_p = vl_bc_icms; aliq_p = aliq_icms; vl_imp_p = vl_icms_i

                # IBS e CBS
                campos_ibs_cbs = (
                    f"{ibs_class}|{ibs_base}|{ibs_aliq}|{ibs_valor}|"
                    f"{cbs_class}|{cbs_base}|{cbs_aliq}|{cbs_valor}"
                )

                # CST de saída para o registro (De-Para)
                pis_cst_saida    = _cst_entrada_para_saida(pis_cst)    if ind_oper == '1' else pis_cst
                cofins_cst_saida = _cst_entrada_para_saida(cofins_cst) if ind_oper == '1' else cofins_cst

                if ind_oper == '0':
                    # ── REGISTRO 1030 (ENTRADA) ───────────────────────────
                    # Erros corrigidos:
                    # - Data: dt_doc_item no formato DD/MM/AAAA (não DDMMAAAA)
                    # - Campo 'Complemento da CFOP' (era 'S') → numérico → ''
                    # - Campo 'Tanque do combustível' (era 'KG'/'UN') → numérico → ''
                    saida.write(
                        f"|1030|{cod_item}|{qtd}|{vl_item}|{vl_ipi}|{vl_bc_icms}|1|{dt_doc_item}||"
                        f"{cod_sit}|{vl_item}|{vl_desc_i}|"
                        f"{vl_bc_icms}|{vl_bc_st}|{aliq_p}|||"
                        f"0,00|0,00|0,00|"
                        f"0,000|{vl_icms_i}|{vl_icms_st}|"
                        f"0,00|0,00|0,00|"
                        f"{vl_unit}|"
                        f"{aliq_st}|{cod_imp}|{aliq_ipi}|"
                        f"0,00|0,00|0,00|"
                        # campo 34=CFOP (numérico) | campo 35=série ECF (vazio)
                        f"{cfop_item}||"
                        # campo 36=aliq_pis | campo 37=vl_pis
                        f"{pis_aliq}|{pis_valor}|"
                        # campo 38=aliq_cofins | campo 39=vl_cofins
                        f"{cofins_aliq}|{cofins_valor}|"
                        # campo 40=custo_total
                        f"0,00|"
                        # campo 41=cst_pis | campo 42=bc_pis
                        f"{pis_cst}|{pis_base}|"
                        # campo 43=cst_cofins | campo 44=bc_cofins
                        f"{cofins_cst}|{cofins_base}|"
                        # campos 45-55 (chassi/lote/arma/enquad_ipi)
                        f"||||||||||{aliq_ipi}|"
                        # campo 56=mov_fisica | campo 57=unid (caractere, OK)
                        # campo 58=complemento CFOP (numérico → vazio)
                        # campo 59=tanque (numérico → vazio)
                        # campo 60=vl_contábil
                        f"S|{unid}|||{vl_item}|"
                        # campos 61-66 = qtd/vlr PIS e COFINS por unidade
                        f"0,000|0,0000|0,00|0,000|0,0000|0,00|"
                        # campo 67=base_credito
                        f"{base_cred_pis}|"
                        # campos 68-69 = nota_dev / descr_comp
                        f"||"
                        # campos 70-71 = CST PIS/COFINS nota devolvida
                        f"||"
                        # campo 72=vinc_pis | campo 73=vinc_cofins
                        f"{vinc_pis}|{vinc_cofins}|"
                        # campos 74-75 = excl PIS/COFINS
                        f"0,00|0,00|"
                        # campos 76-78 = ICMS carga média
                        f"0,00|0,00|0,00|"
                        # campos 79-84 = ECF/redução/cod_rec
                        f"||0,00|||{nat_pis}|{nat_cofins}|"
                        # campos 85-89 = crédito presumido + ICMS ST antecipação
                        f"0,00|0,00|0,00|0,00|0,00|"
                        # campos 90-94 = cod_rec_ipi/cest/ICMS_ST_retido
                        f"||0,00|0,00||"
                        # campo 95 = CBENEF (Identificador)
                        f"{cbenef_item}|"
                        # campos 96-103
                        f"0,00|0,00|||0,0000|0,0000|0,0000||"
                        # campos 104-111 = IBS e CBS
                        f"{campos_ibs_cbs}||||\n"
                    )
                else:
                    # ── REGISTRO 2030 (SAÍDA) ─────────────────────────────
                    saida.write(
                        f"|2030|{cod_item}|{qtd}|{vl_item}|{vl_ipi}|{vl_bc_icms}|1|{dt_doc_item}|"
                        f"{cod_sit}|{vl_item}|{vl_desc_i}|"
                        f"{vl_bc_icms}|{vl_bc_st}|{aliq_p}|"
                        f"||||"
                        f"0,00|0,00|0,00|"
                        f"0,000|{vl_icms_i}|{vl_icms_st}|"
                        f"0,00|0,00|"
                        f"{vl_unit}|{aliq_st}|{cod_imp}|{aliq_ipi}|"
                        f"0,00|0,00|0,00|"
                        f"|||||||0,00|0,00|0,00|0,00|0,00|||"
                        # campos 48-55 = PIS e COFINS (CST de saída)
                        f"{pis_cst_saida}|{pis_base}|{pis_aliq}|{pis_valor}|"
                        f"{cofins_cst_saida}|{cofins_base}|{cofins_aliq}|{cofins_valor}|"
                        # campos 56-62
                        f"||{aliq_ipi}|{qtd}|||{vl_item}|"
                        # campos 63-68
                        f"0,000|0,0000|0,00|0,000|0,0000|0,00|"
                        # campo 69=nota_dev | campo 70=descr_comp
                        f"||"
                        # campo 71=nat_pis | campo 72=nat_cofins
                        f"{nat_pis}|{nat_cofins}|"
                        # campo 73=excl_coop
                        f"0,00|"
                        # campos 74-75 = CST dev
                        f"||"
                        # campo 76=data | campo 77=vinc_pis | campo 78=vinc_cofins
                        f"{dt_doc_item}|{vinc_pis}|{vinc_cofins}|"
                        # campos 79-90
                        f"0,00|0,00|0,00|0,00|0,00|0,00|0,00|0,00|0,00|0,00|0,00|0,00|"
                        # campos 91-95
                        f"|0,00|0,00|0,00||"
                        # campos 96-100 = DIFAL
                        f"0,00|0,00|0,00|0,00|0,00|"
                        # campos 101-106
                        f"||||0,00||"
                        # campo 107 = CBENEF
                        f"{cbenef_item}|"
                        # campos 108-110
                        f"||0,00||"
                        # campos 111-118 = IBS e CBS
                        f"{campos_ibs_cbs}||||\n"
                    )

                stats['itens'] += 1

        except Exception as e:
            log.append(f"ERRO ao converter C100 NF={_c(campos_c100, SPED_C100_NUM_DOC)}: {e}")
            log.append(traceback.format_exc())
            stats['erros'] += 1

    # D100
    if 'D100' in parsed['por_tipo']:
        for campos, num_linha in parsed['por_tipo']['D100']:
            try:
                ind_oper = _c(campos, SPED_D100_IND_OPER)
                cod_mod  = _c(campos, SPED_D100_COD_MOD)
                cod_sit  = _c(campos, SPED_D100_COD_SIT)
                serie    = _c(campos, SPED_D100_SER)
                num_doc  = _c(campos, SPED_D100_NUM_DOC)
                dt_doc   = converter_data(_c(campos, SPED_D100_DT_DOC))
                vl_doc   = _c(campos, SPED_D100_VL_DOC)
                aliq     = _c(campos, SPED_D100_ALIQ)
                vl_icms  = _c(campos, SPED_D100_VL_ICMS)
                tipo_es  = 'E' if ind_oper == '0' else 'S'
                cod_part = _c(campos, SPED_D100_COD_PART)
                part_campos = participantes.get(cod_part, [])
                cnpj_part   = _c(part_campos, SPED_0150_CNPJ) if part_campos else ''
                saida.write(
                    f"|1000|{num_doc}|{cnpj_part}||{ind_oper}||"
                    f"{serie}|{cod_mod}|{cod_sit}||||"
                    f"{dt_doc}|{dt_doc}|{vl_doc}||FRETE||||||||||{tipo_es}|"
                    f"0,00|0,00|0,00|0,00||0,00||||0,00|0,00|0,00||{vl_doc}|"
                    f"0|0|||||0,00||||||N|S||{tipo_es}||0|||||"
                    f"||||||||||||0|{cod_sit}|0||0,00|0,00|0,00|||||||||||\n"
                )
                saida.write(
                    f"|1020|{num_doc}||{vl_doc}|{aliq}|{vl_icms}|"
                    f"0,00|0,00|0,00|0,00|{vl_doc}||||\n"
                )
                stats['transporte'] += 1
            except Exception as e:
                log.append(f"ERRO D100 linha {num_linha}: {e}"); stats['erros'] += 1

    # H010
    if 'H010' in parsed['por_tipo']:
        for campos, _ in parsed['por_tipo']['H010']:
            saida.write(
                f"|H010|{_c(campos, SPED_H010_COD_ITEM)}|{_c(campos, SPED_H010_UNID)}|"
                f"{_c(campos, SPED_H010_QTD)}|{_c(campos, SPED_H010_VL_UNIT)}|"
                f"{_c(campos, SPED_H010_VL_ITEM)}|\n"
            )
            stats['inventario'] += 1

    saida.write("|9999|\n")

    if nao_mapeados:
        log.append(f"AVISO: {len(nao_mapeados)} CFOP(s) sem acumulador (9999): "
                   f"{', '.join(sorted(nao_mapeados))}")
    log.append(
        f"Conversão concluída — Produtos={stats['produtos']} | "
        f"NFs entrada={stats['nf_entrada']} | NFs saída={stats['nf_saida']} | "
        f"Itens={stats['itens']} | Da Planilha={stats['itens_planilha']} | "
        f"Do SPED={stats['itens_sped']} | Natureza={stats['natureza_aplicada']} | "
        f"CBENEF={stats['cbenef_aplicado']} | Devoluções={stats['devolucoes']} | "
        f"Erros={stats['erros']}"
    )
    return saida.getvalue(), stats


# ==============================
# INTERFACE STREAMLIT — V3.3
# Upload SPED e Planilha Cliente lado a lado
# ==============================
def main():
    st.set_page_config(
        page_title="SPED Fiscal → Domínio | Thomson Reuters",
        page_icon="🟠", layout="wide", initial_sidebar_state="expanded",
    )
    apply_tr_theme()

    tabela_cfop, tabela_flags    = carregar_tabela_cfop_oficial()
    por_codigo_imp, por_nome_imp = carregar_tabela_impostos()

    st.markdown(
        f"""
        <div style="background:#444444; padding:24px 28px 18px 28px; border-radius:8px;
                    border-top:6px solid #FF8000; margin-bottom:28px;">
            <h2 style="color:#FF8000; margin:0; font-family:'Segoe UI',Arial,sans-serif;">
                📄 Conversor SPED Fiscal → Domínio Sistemas &nbsp;|&nbsp; {VERSAO}
            </h2>
            <p style="color:#DDDDDD; margin:6px 0 0 0; font-family:'Segoe UI',Arial,sans-serif;">
                Entrada: <strong>SPED Fiscal EFD ICMS/IPI</strong> &nbsp;+&nbsp;
                <strong>Planilha Cliente</strong> &nbsp;→&nbsp;
                Saída: <strong>Leiaute padrão Domínio Sistemas</strong>
                &nbsp;|&nbsp; CFOPs: <strong>{len(tabela_cfop)}</strong>
                &nbsp;|&nbsp; Impostos: <strong>{len(por_codigo_imp)}</strong>
            </p>
        </div>
        """, unsafe_allow_html=True,
    )

    with st.sidebar:
        st.markdown("### ℹ Sobre")
        st.markdown(f"**Versão:** {VERSAO}")
        st.markdown("**Thomson Reuters  |  Domínio Sistemas**")
        st.markdown("---")
        st.markdown("### 🔄 De-Para CST Entrada → Saída")
        st.markdown(
            "| CST Entrada | CST Saída |\n|---|---|\n"
            "| 50 | 01 – Alíq. Básica |\n"
            "| 70 | 07 – Isenta |\n"
            "| 72 | 09 – Suspensão |\n"
            "| 73 | 06 – Alíq. Zero |\n"
            "| 74 | 08 – Sem Incidência |\n"
            "| 75 | 05 – Subst. Tributária |\n\n"
            "Usado no campo 16 do **0110** (CST Saídas)."
        )
        st.markdown("---")
        st.markdown("### 🏷️ Natureza de Receita")
        st.markdown(
            "Exibida **apenas** para CSTs obrigatórios:\n"
            "`06`, `07`, `08`, `09`\n\n"
            "CSTs opcionais **não aparecem**."
        )
        st.markdown("---")
        st.markdown("### 🔵 CBENEF")
        st.markdown(
            "Coluna `CBENEF` da Planilha Cliente.\n\n"
            "Campos:\n"
            "- `0100` campo **91** (Identificador)\n"
            "- `1030` campo **95**\n"
            "- `2030` campo **107**"
        )
        st.markdown("---")
        st.markdown("### 📑 Fluxo")
        st.markdown(
            "1. **Upload simultâneo** SPED + Planilha Cliente\n"
            "2. Extrair CFOPs → baixar XLSX\n"
            "3. Preencher `ACUMULADOR`\n"
            "4. Configurar Natureza (só CSTs obrigatórios)\n"
            "5. Upload XLSX de acumuladores\n"
            "6. Converter e baixar saída\n"
        )
        st.markdown("---")
        if tabela_cfop:
            st.success(f"✅ {len(tabela_cfop)} CFOPs carregados")
        else: st.error("❌ Tabela CFOP não carregada!")
        if por_codigo_imp: st.success(f"✅ {len(por_codigo_imp)} impostos carregados")
        else: st.warning("⚠ Tabela de impostos não carregada (fallback).")

    st.markdown("---")

    defaults = {
        "log":                    [f"Aplicação pronta. {VERSAO} | CFOPs: {len(tabela_cfop)} | Impostos: {len(por_codigo_imp)}"],
        "resultado":              None, "nome_saida": "saida_dominio.txt",
        "stats":                  None, "xlsx_bytes": None, "xlsx_nome": "acumuladores.xlsx",
        "cfops_extraidos":        None, "tabela_acum_ok": False,
        "arquivo_raw":            None, "arquivo_nome": None,
        "df_cliente":             None, "cliente_ok": False,
        "config_natureza":        {}, "cbenef_por_produto": {},
        "csts_obrigatorios":      set(),
        "info_produto_planilha":  {},
    }
    for k, v in defaults.items():
        if k not in st.session_state: st.session_state[k] = v

    # ── Etapa 1: Upload SPED + Planilha Cliente (lado a lado) ─────────────
    st.markdown("### 📥 Etapa 1 — Upload dos Arquivos")
    st.markdown("""
    <div class="cbenef-box">
    Faça o upload do <strong>SPED Fiscal</strong> e da <strong>Planilha Cliente</strong>
    simultaneamente. Os dados de PIS, COFINS, IBS, CBS e CBENEF da planilha serão
    automaticamente combinados com as informações do SPED para gerar o arquivo Domínio.
    </div>
    """, unsafe_allow_html=True)

    col_sped, col_cliente = st.columns(2)

    with col_sped:
        st.markdown("#### 📄 SPED Fiscal")
        uploaded_sped = st.file_uploader(
            "Arquivo SPED Fiscal (.txt)", type=["txt"], key="upload_sped",
            help="Arquivo EFD ICMS/IPI exportado pelo ERP"
        )
        if uploaded_sped is not None:
            raw_atual = uploaded_sped.read()
            if raw_atual != st.session_state.arquivo_raw:
                st.session_state.arquivo_raw     = raw_atual
                st.session_state.arquivo_nome    = uploaded_sped.name
                st.session_state.cfops_extraidos = None
                st.session_state.xlsx_bytes      = None
                st.session_state.resultado       = None
                st.session_state.stats           = None
                st.session_state.tabela_acum_ok  = False
                st.session_state.log             = [
                    f"SPED carregado: {uploaded_sped.name} ({len(raw_atual)/1024:.1f} KB)"
                ]
        if st.session_state.arquivo_raw is not None:
            st.success(f"✅ **{st.session_state.arquivo_nome}** "
                       f"({len(st.session_state.arquivo_raw)/1024:.1f} KB)")

    with col_cliente:
        st.markdown("#### 📊 Planilha Cliente")
        uploaded_cliente = st.file_uploader(
            "Planilha Cliente (.xlsx ou .csv)", type=["xlsx", "xls", "csv"],
            key="upload_cliente",
            help="Planilha com PIS CST, COFINS CST, alíquotas, bases, IBS, CBS e CBENEF"
        )
        if uploaded_cliente is not None:
            log_temp = []; raw_cli = uploaded_cliente.read()
            df_cli_prev = carregar_planilha_cliente(raw_cli, uploaded_cliente.name, log_temp)
            if df_cli_prev is not None:
                st.session_state.df_cliente = df_cli_prev
                st.session_state.cliente_ok = True
                # Extrai informações
                csts_obrig   = extrair_csts_obrigatorios_da_planilha(df_cli_prev, log_temp)
                cb_map_prev  = extrair_cbenef_por_produto(df_cli_prev, log_temp)
                info_pl_prev = extrair_info_produto_da_planilha(df_cli_prev, log_temp)
                st.session_state.csts_obrigatorios     = csts_obrig
                st.session_state.cbenef_por_produto    = cb_map_prev
                st.session_state.info_produto_planilha = info_pl_prev
                n_cb = len([v for v in cb_map_prev.values() if v])

                # CSTs encontrados (todos) para exibição informativa
                todos_csts_raw = set()
                for col in ('PIS CST', 'COFINS CST'):
                    if col in df_cli_prev.columns:
                        for val in df_cli_prev[col].dropna():
                            s = str(val).strip().split('.')[0]
                            try: todos_csts_raw.add(str(int(s)).zfill(2))
                            except ValueError:
                                if s: todos_csts_raw.add(s.zfill(2))

                csts_opcio = todos_csts_raw - CSTS_NATUREZA_OBRIGATORIA

                st.success(
                    f"✅ **{uploaded_cliente.name}** — "
                    f"{len(df_cli_prev)} linhas | "
                    f"NFs: {df_cli_prev['NF'].nunique()} | "
                    f"Itens: {df_cli_prev['COD.ITEM'].nunique()}"
                )
                # Exibe CSTs detectados
                cst_info_parts = []
                if csts_obrig:
                    cst_info_parts.append(
                        f"⚠ Obrigatórios: " +
                        ", ".join([f"`{c}`" for c in sorted(csts_obrig)])
                    )
                if csts_opcio:
                    cst_info_parts.append(
                        f"ℹ Opcionais: " +
                        ", ".join([f"`{c}`" for c in sorted(csts_opcio)])
                    )
                if cst_info_parts:
                    st.markdown(" &nbsp;·&nbsp; ".join(cst_info_parts))
                if n_cb > 0:
                    st.markdown(f"🔵 CBENEF: **{n_cb}** produto(s)")
                if log_temp:
                    st.session_state.log.extend(log_temp)
            else:
                st.session_state.df_cliente = None
                st.session_state.cliente_ok = False
                st.session_state.csts_obrigatorios = set()
                st.session_state.cbenef_por_produto = {}
                st.session_state.info_produto_planilha = {}
                for msg in log_temp: st.error(msg)
        else:
            if not st.session_state.cliente_ok:
                st.info("⬆ Upload opcional. Sem a planilha, usa dados do SPED como fallback.")

    st.markdown("---")

    # ── Extração de CFOPs ─────────────────────────────────────────────────
    st.markdown("### 🔍 Etapa 2 — Extrair CFOPs e Gerar Planilha de Acumuladores")

    col_e1, col_e2 = st.columns(2)
    with col_e1:
        extrair = st.button("🔍 Extrair CFOPs e Gerar Planilha",
                            disabled=(st.session_state.arquivo_raw is None),
                            use_container_width=True, type="primary")
    with col_e2:
        if st.session_state.xlsx_bytes is not None:
            st.download_button(
                label="⬇ Baixar Planilha de Acumuladores (.xlsx)",
                data=st.session_state.xlsx_bytes, file_name=st.session_state.xlsx_nome,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary",
            )

    if extrair:
        st.session_state.log = [f"Extraindo CFOPs do SPED Fiscal — {st.session_state.arquivo_nome}..."]
        st.session_state.xlsx_bytes = None; st.session_state.cfops_extraidos = None
        try:
            content    = decode_arquivo(st.session_state.arquivo_raw)
            parsed     = parse_sped(content)
            st.session_state.log.append(f"Registros: {', '.join(parsed['por_tipo'].keys())}")
            cfops_dict = extrair_cfops_do_sped(parsed, tabela_flags, st.session_state.log)
            if not cfops_dict:
                st.session_state.log.append("AVISO: Nenhum CFOP encontrado.")
            else:
                xlsx_bytes = gerar_xlsx_acumuladores_tr(cfops_dict, tabela_cfop, tabela_flags)
                nome_base  = st.session_state.arquivo_nome.replace('.txt', '')
                nome_xlsx  = f"{nome_base}_acumuladores.xlsx"
                st.session_state.xlsx_bytes      = xlsx_bytes
                st.session_state.xlsx_nome       = nome_xlsx
                st.session_state.cfops_extraidos = cfops_dict
                n_dev = sum(1 for v in cfops_dict.values() if v.get('indDevol') == 1)
                st.session_state.log.append(
                    f"✔ {len(cfops_dict)} CFOP(s) | {n_dev} devolução(ões) | {nome_xlsx}"
                )
        except Exception:
            st.session_state.log.append("ERRO FATAL na extração de CFOPs.")
            st.session_state.log.append(traceback.format_exc())
        st.rerun()

    if st.session_state.cfops_extraidos:
        cfops_dict = st.session_state.cfops_extraidos
        n_ent = sum(1 for v in cfops_dict.values() if v['tipo_operacao'] == 'Entrada')
        n_sai = sum(1 for v in cfops_dict.values() if v['tipo_operacao'] == 'Saída')
        n_dev = sum(1 for v in cfops_dict.values() if v.get('indDevol') == 1)
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        col_m1.metric("CFOPs únicos", len(cfops_dict))
        col_m2.metric("Entradas", n_ent)
        col_m3.metric("Saídas", n_sai)
        col_m4.metric("Devoluções", n_dev)

    st.markdown("---")

    # ── Etapa 3: Natureza (apenas CSTs obrigatórios da planilha) ──────────
    csts_para_natureza = st.session_state.get('csts_obrigatorios', set())
    config_natureza = render_configuracao_natureza(csts_para_natureza)

    st.markdown("---")

    # ── Etapa 4: Acumuladores + Conversão ─────────────────────────────────
    st.markdown("### ▶ Etapa 3 — Converter com a tabela de acumuladores preenchida")
    arquivo_acum = st.file_uploader(
        "📂 Tabela de Acumuladores preenchida (.xlsx ou .csv)",
        type=["xlsx", "xls", "csv"], key="upload_acum",
    )
    if arquivo_acum is not None:
        log_temp = []; raw_acum = arquivo_acum.read()
        tab_prev = carregar_acumuladores(raw_acum, arquivo_acum.name, log_temp)
        arquivo_acum.seek(0)
        if tab_prev is not None:
            st.success(f"✅ Tabela válida — **{len(tab_prev)} CFOPs** com acumulador preenchido.")
            st.session_state.tabela_acum_ok = True
        else:
            for msg in log_temp: st.error(msg)
            st.session_state.tabela_acum_ok = False
    else:
        if not st.session_state.tabela_acum_ok:
            st.info("⬆ Faça o upload da tabela de acumuladores preenchida para converter.")

    pode_converter = st.session_state.tabela_acum_ok and st.session_state.arquivo_raw is not None
    col1, col2 = st.columns(2)
    with col1:
        converter = st.button("▶ Converter SPED → Domínio",
                              disabled=not pode_converter,
                              use_container_width=True, type="primary")
    with col2:
        limpar = st.button("🗑 Limpar Tudo", use_container_width=True)

    if limpar:
        for k in list(st.session_state.keys()): del st.session_state[k]
        st.rerun()

    if converter and pode_converter:
        st.session_state.log       = ["Iniciando conversão SPED → Domínio Sistemas..."]
        st.session_state.resultado = None; st.session_state.stats = None
        try:
            arquivo_acum.seek(0)
            tabela_acum = carregar_acumuladores(
                arquivo_acum.read(), arquivo_acum.name, st.session_state.log)
            if tabela_acum is None:
                st.session_state.log.append("ERRO: Tabela inválida. Abortando.")
                st.rerun()
            content = decode_arquivo(st.session_state.arquivo_raw)
            parsed  = parse_sped(content)
            resultado_txt, stats = converter_sped_para_dominio(
                parsed, tabela_acum, tabela_flags, por_nome_imp,
                st.session_state.log,
                df_cliente=st.session_state.get('df_cliente'),
                config_natureza=st.session_state.get('config_natureza', {}),
                cbenef_por_produto=st.session_state.get('cbenef_por_produto', {}),
                info_produto_planilha=st.session_state.get('info_produto_planilha', {}),
            )
            resultado_bytes = encode_ansi_seguro(resultado_txt, st.session_state.log)
            st.session_state.resultado  = resultado_bytes
            st.session_state.stats      = stats
            st.session_state.nome_saida = st.session_state.arquivo_nome.replace('.txt', '_dominio.txt')
        except Exception:
            st.session_state.log.append("ERRO FATAL durante a conversão.")
            st.session_state.log.append(traceback.format_exc())
        st.rerun()

    if st.session_state.resultado is not None:
        st.success("✅ Arquivo convertido com sucesso!")
        stats = st.session_state.stats or {}
        st.markdown("#### 📊 Estatísticas da Conversão")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Produtos",            stats.get('produtos',         0))
        col2.metric("NFs Entrada",         stats.get('nf_entrada',       0))
        col3.metric("NFs Saída",           stats.get('nf_saida',         0))
        col4.metric("Devoluções",          stats.get('devolucoes',        0))
        col5, col6, col7, col8 = st.columns(4)
        col5.metric("Itens (total)",       stats.get('itens',             0))
        col6.metric("Da Planilha Cliente", stats.get('itens_planilha',    0))
        col7.metric("Natureza Aplicada",   stats.get('natureza_aplicada', 0))
        col8.metric("CBENEF Aplicado",     stats.get('cbenef_aplicado',   0))
        col9, _ = st.columns([1, 3])
        col9.metric("Erros", stats.get('erros', 0))
        with st.expander("👁️ Prévia do arquivo gerado (primeiras 80 linhas)"):
            preview = '\n'.join(
                st.session_state.resultado.decode('latin-1', errors='replace').splitlines()[:80]
            )
            st.code(preview, language='text')
        st.download_button(
            label="⬇ Baixar Arquivo Domínio Sistemas",
            data=st.session_state.resultado,
            file_name=st.session_state.nome_saida,
            mime="text/plain", use_container_width=True, type="primary",
        )

    st.markdown("**Log de processamento**")
    log_texto = "\n".join(st.session_state.log)
    tem_erro  = any(str(l).startswith("ERRO") for l in st.session_state.log)
    cor_borda = "#D32F2F" if tem_erro else "#388E3C"
    st.markdown(
        f"""<div style="background:#FCFCFC; border:1px solid {cor_borda}; border-radius:6px;
                    padding:14px; font-family:Consolas,monospace; font-size:13px;
                    white-space:pre-wrap; max-height:340px; overflow-y:auto; color:#1F1F1F;">
{log_texto}</div>""", unsafe_allow_html=True,
    )
    st.markdown("---")
    st.caption("Conversor SPED Fiscal → Domínio Sistemas | Thomson Reuters | Python + Streamlit")


if __name__ == "__main__":
    main()
